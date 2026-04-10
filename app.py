import streamlit as st
import sqlite3
import pandas as pd
import os
import base64
import re
import json
import subprocess
import sys
from pathlib import Path
from PIL import Image
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import io
import streamlit.components.v1 as components

# ── Módulo de persistência ──────────────────────────────────────────────────
import database as db_module

# ── Módulo de revisão colaborativa ───────────────────────────────────────────
try:
    import ai_review
    _AI_REVIEW_OK = True
except ImportError:
    _AI_REVIEW_OK = False

# -- Módulo de página de Diário de Auditoria --
try:
    import plotly.express as px
    import plotly.graph_objects as go
    _PLOTLY_OK = True
except ImportError:
    _PLOTLY_OK = False

# ──────────────────────────────────────────────────────────────────────────────
# SELETOR DE ARQUIVOS/PASTAS (Explorador Windows nativo via tkinter)
# ──────────────────────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────────────────
# LIGHTBOX: galeria de imagens com clique direto e setas de navegação
# ─────────────────────────────────────────────────────────────────────────────
def render_image_gallery_html(imgs: list, key_prefix: str):
    """Renderiza galeria paginada com lightbox HTML/JS.

    - Sem limite de imagens (exibe TODAS)
    - Paginação JavaScript (20 por página) – sem rerun do Streamlit
    - Thumb 200×200, Full 900×700 → HTML leve mesmo com 100+ imagens
    - Lightbox com setas, teclado e clique fora para fechar
    """
    import base64, io as _io, json as _json
    from PIL import Image as _Image

    THUMB_MAX  = (200, 200)
    FULL_MAX   = (900, 700)
    PAGE_SIZE  = 20           # imagens por página na grade

    b64_thumbs = []
    b64_fulls  = []
    names_list = []

    for img_path in imgs:
        try:
            with _Image.open(img_path) as im:
                fmt = im.format or "JPEG"
                if fmt not in ("JPEG","PNG","GIF","WEBP","BMP"):
                    fmt = "JPEG"
                # --- thumb ---
                th = im.copy()
                th.thumbnail(THUMB_MAX)
                buf_t = _io.BytesIO()
                th.save(buf_t, format=fmt, quality=70, optimize=True)
                b64_thumbs.append(base64.b64encode(buf_t.getvalue()).decode())
                # --- full ---
                fl = im.copy()
                fl.thumbnail(FULL_MAX)
                buf_f = _io.BytesIO()
                fl.save(buf_f, format=fmt, quality=80, optimize=True)
                b64_fulls.append(base64.b64encode(buf_f.getvalue()).decode())
            names_list.append(img_path.name)
        except Exception:
            b64_thumbs.append("")
            b64_fulls.append("")
            names_list.append(getattr(img_path, "name", str(img_path)))

    n   = len(names_list)
    uid = key_prefix.replace("-","_").replace(".","_")

    thumbs_json = _json.dumps(b64_thumbs)
    fulls_json  = _json.dumps(b64_fulls)
    names_json  = _json.dumps(names_list)
    page_size_js = PAGE_SIZE

    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
  body {{ margin:0; padding:0; background:transparent; font-family:sans-serif; }}

  /* ── Grade ── */
  .lb-grid {{
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 6px;
    padding: 4px;
  }}
  .lb-thumb-wrap {{ display:flex; flex-direction:column; }}
  .lb-thumb {{
    cursor:pointer; border-radius:5px; overflow:hidden;
    border:2px solid transparent; transition:border-color .2s, transform .15s;
    background:#f0f0f0; aspect-ratio:1/1;
    display:flex; align-items:center; justify-content:center;
  }}
  .lb-thumb:hover {{ border-color:#3D5AFE; transform:scale(1.04); }}
  .lb-thumb img {{ width:100%; height:100%; object-fit:cover; display:block; }}
  .lb-caption {{
    font-size:9px; color:#555; text-align:center;
    overflow:hidden; text-overflow:ellipsis; white-space:nowrap;
    padding:2px 2px 0; max-width:100%;
  }}

  /* ── Paginação ── */
  .lb-pag {{
    display:flex; align-items:center; justify-content:center;
    gap:8px; margin:10px 0 4px;
    font-size:13px; color:#444;
  }}
  .lb-pag-btn {{
    background:#e8eaed; border:none; border-radius:5px;
    padding:5px 14px; cursor:pointer; font-size:13px; font-weight:600;
    transition:background .18s;
  }}
  .lb-pag-btn:hover {{ background:#c5cae9; }}
  .lb-pag-btn:disabled {{ opacity:.4; cursor:not-allowed; }}
  .lb-pag-info {{ font-weight:600; color:#3D5AFE; min-width:90px; text-align:center; }}

  /* ── Lightbox overlay ── */
  #lb-ov-{uid} {{
    display:none; position:fixed; inset:0; z-index:999999;
    background:rgba(0,0,0,.88); align-items:center;
    justify-content:center; flex-direction:column;
  }}
  #lb-ov-{uid}.lb-active {{ display:flex; }}
  #lb-img-{uid} {{
    max-width:88vw; max-height:80vh; border-radius:8px;
    box-shadow:0 8px 40px #000a; object-fit:contain; user-select:none;
  }}
  .lb-nav-{uid} {{
    display:flex; align-items:center; gap:16px; margin-top:14px;
  }}
  .lb-nbtn-{uid} {{
    background:#fff2; color:#fff; border:none; padding:10px 22px;
    border-radius:6px; cursor:pointer; font-size:18px; font-weight:bold;
    transition:background .18s;
  }}
  .lb-nbtn-{uid}:hover {{ background:#fff4; }}
  .lb-nbtn-{uid}:disabled {{ opacity:.35; cursor:not-allowed; }}
  .lb-cls-{uid} {{
    position:absolute; top:18px; right:24px; background:transparent;
    border:none; color:#fff; font-size:30px; cursor:pointer; line-height:1;
  }}
  .lb-ctr-{uid} {{ color:#ddd; font-size:14px; min-width:80px; text-align:center; }}
  .lb-nm-{uid} {{
    color:#aaa; font-size:12px; margin-top:6px; max-width:80vw;
    overflow:hidden; text-overflow:ellipsis; white-space:nowrap; text-align:center;
  }}
</style>
</head>
<body>

<div id="grid-{uid}" class="lb-grid"></div>
<div class="lb-pag">
  <button class="lb-pag-btn" id="pbtn-prev-{uid}">&#8592; Anterior</button>
  <span  class="lb-pag-info" id="pbtn-info-{uid}">Pág 1 / 1</span>
  <button class="lb-pag-btn" id="pbtn-next-{uid}">Próxima &#8594;</button>
  <span style="color:#888;font-size:12px" id="pbtn-total-{uid}"></span>
</div>

<script>
(function(){{
  var THUMBS    = {thumbs_json};
  var FULLS     = {fulls_json};
  var NAMES     = {names_json};
  var N         = NAMES.length;
  var PAGE_SIZE = {page_size_js};
  var uid       = "{uid}";
  var curPage   = 0;
  var curImg    = 0;
  var totalPages= Math.ceil(N / PAGE_SIZE);

  var parentDoc = window.parent ? window.parent.document : document;

  // ── Injetar overlay lightbox no parent uma única vez ──
  if (!parentDoc.getElementById("lb-ov-" + uid)) {{
    var ov  = parentDoc.createElement("div");
    ov.id   = "lb-ov-" + uid;
    ov.innerHTML = `
      <button class="lb-cls-{uid}" id="lb-cls-btn-{uid}" title="Fechar">&times;</button>
      <img id="lb-img-{uid}" src="" alt="">
      <div class="lb-nm-{uid}"  id="lb-nm-{uid}"></div>
      <div class="lb-nav-{uid}">
        <button class="lb-nbtn-{uid}" id="lb-prev-{uid}">&#8592; Anterior</button>
        <span   class="lb-ctr-{uid}" id="lb-ctr-{uid}">1 / 1</span>
        <button class="lb-nbtn-{uid}" id="lb-next-{uid}">Próxima &#8594;</button>
      </div>
    `;
    var sty = parentDoc.createElement("style");
    sty.textContent = `
      #lb-ov-{uid}{{display:none;position:fixed;inset:0;z-index:999999;
        background:rgba(0,0,0,.88);align-items:center;justify-content:center;
        flex-direction:column;}}
      #lb-ov-{uid}.lb-active{{display:flex;}}
      #lb-img-{uid}{{max-width:88vw;max-height:80vh;border-radius:8px;
        box-shadow:0 8px 40px #000a;object-fit:contain;user-select:none;}}
      .lb-nav-{uid}{{display:flex;align-items:center;gap:16px;margin-top:14px;}}
      .lb-nbtn-{uid}{{background:#fff2;color:#fff;border:none;padding:10px 22px;
        border-radius:6px;cursor:pointer;font-size:18px;font-weight:bold;
        transition:background .18s;}}
      .lb-nbtn-{uid}:hover{{background:#fff4;}}
      .lb-nbtn-{uid}:disabled{{opacity:.35;cursor:not-allowed;}}
      .lb-cls-{uid}{{position:absolute;top:18px;right:24px;background:transparent;
        border:none;color:#fff;font-size:30px;cursor:pointer;line-height:1;}}
      .lb-ctr-{uid}{{color:#ddd;font-size:14px;min-width:80px;text-align:center;}}
      .lb-nm-{uid}{{color:#aaa;font-size:12px;margin-top:6px;max-width:80vw;
        overflow:hidden;text-overflow:ellipsis;white-space:nowrap;text-align:center;}}
    `;
    parentDoc.head.appendChild(sty);
    parentDoc.body.appendChild(ov);

    parentDoc.getElementById("lb-prev-" + uid).onclick = function() {{
      if (curImg > 0) showImg(curImg - 1);
    }};
    parentDoc.getElementById("lb-next-" + uid).onclick = function() {{
      if (curImg < N - 1) showImg(curImg + 1);
    }};
    parentDoc.getElementById("lb-cls-btn-" + uid).onclick = function() {{
      parentDoc.getElementById("lb-ov-" + uid).classList.remove("lb-active");
    }};
    parentDoc.getElementById("lb-ov-" + uid).addEventListener("click", function(e) {{
      if (e.target === this) this.classList.remove("lb-active");
    }});
    parentDoc.addEventListener("keydown", function(e) {{
      var ov2 = parentDoc.getElementById("lb-ov-" + uid);
      if (!ov2 || !ov2.classList.contains("lb-active")) return;
      if (e.key === "ArrowLeft"  && curImg > 0)   showImg(curImg - 1);
      if (e.key === "ArrowRight" && curImg < N-1) showImg(curImg + 1);
      if (e.key === "Escape") ov2.classList.remove("lb-active");
    }});
  }}

  // ── Lightbox: mostrar imagem ──
  function showImg(idx) {{
    curImg = idx;
    var im  = parentDoc.getElementById("lb-img-" + uid);
    var nm  = parentDoc.getElementById("lb-nm-"  + uid);
    var ct  = parentDoc.getElementById("lb-ctr-" + uid);
    var prv = parentDoc.getElementById("lb-prev-" + uid);
    var nxt = parentDoc.getElementById("lb-next-" + uid);
    im.src  = "data:image/jpeg;base64," + FULLS[idx];
    nm.textContent = NAMES[idx];
    ct.textContent = (idx+1) + " / " + N;
    prv.disabled   = (idx === 0);
    nxt.disabled   = (idx === N - 1);
  }}

  function openLightbox(idx) {{
    parentDoc.getElementById("lb-ov-" + uid).classList.add("lb-active");
    showImg(idx);
  }}

  // ── Renderizar página ──
  function renderPage(page) {{
    curPage = page;
    var grid  = document.getElementById("grid-" + uid);
    var start = page * PAGE_SIZE;
    var end   = Math.min(start + PAGE_SIZE, N);
    grid.innerHTML = "";
    for (var i = start; i < end; i++) {{
      (function(idx) {{
        var wrap = document.createElement("div");
        wrap.className = "lb-thumb-wrap";
        var cell = document.createElement("div");
        cell.className = "lb-thumb";
        var img  = document.createElement("img");
        img.src  = "data:image/jpeg;base64," + THUMBS[idx];
        img.alt  = NAMES[idx];
        cell.appendChild(img);
        cell.onclick = function() {{ openLightbox(idx); }};
        var cap = document.createElement("div");
        cap.className  = "lb-caption";
        cap.textContent = NAMES[idx];
        wrap.appendChild(cell);
        wrap.appendChild(cap);
        grid.appendChild(wrap);
      }})(i);
    }}
    // Paginação
    var info  = document.getElementById("pbtn-info-" + uid);
    var total = document.getElementById("pbtn-total-" + uid);
    var prev  = document.getElementById("pbtn-prev-" + uid);
    var next  = document.getElementById("pbtn-next-" + uid);
    info.textContent  = "Pág " + (page+1) + " / " + totalPages;
    total.textContent = N + " imagens";
    prev.disabled     = (page === 0);
    next.disabled     = (page >= totalPages - 1);
  }}

  document.getElementById("pbtn-prev-" + uid).onclick = function() {{
    if (curPage > 0) renderPage(curPage - 1);
  }};
  document.getElementById("pbtn-next-" + uid).onclick = function() {{
    if (curPage < totalPages - 1) renderPage(curPage + 1);
  }};

  // ── Init ──
  renderPage(0);
}})();
</script>
</body>
</html>
"""
    # Altura: 4 colunas, cada célula ~110px + paginação 40px
    rows_per_page = (min(PAGE_SIZE, n) + 3) // 4
    height = max(160, rows_per_page * 110 + 60)
    components.html(html, height=height, scrolling=True)




def _run_tkinter_dialog(script: str) -> str:
    """Executa um pequeno script tkinter em subprocesso e retorna o caminho selecionado."""
    try:
        result = subprocess.run(
            [sys.executable, '-c', script],
            capture_output=True, text=True, timeout=60
        )
        return result.stdout.strip()
    except Exception:
        return ''

def browser_pasta(titulo: str = "Selecionar Pasta", initial_dir: str = "") -> str:
    """Abre o Explorador de Arquivos do Windows para selecionar uma PASTA."""
    ini = initial_dir.replace("\\", "/") if initial_dir else ""
    ini_arg = f', initialdir="{ini}"' if ini else ''
    script = f'''
import tkinter as tk
from tkinter import filedialog
root = tk.Tk()
root.withdraw()
root.wm_attributes('-topmost', 1)
folder = filedialog.askdirectory(title="{titulo}"{ini_arg})
print(folder)
'''
    return _run_tkinter_dialog(script)

def browser_arquivo_xlsx(titulo: str = "Selecionar Arquivo Excel", initial_dir: str = "") -> str:
    """Abre o Explorador de Arquivos do Windows para selecionar um arquivo .xlsx."""
    ini = initial_dir.replace("\\", "/") if initial_dir else ""
    ini_arg = f', initialdir="{ini}"' if ini else ''
    script = f'''
import tkinter as tk
from tkinter import filedialog
root = tk.Tk()
root.withdraw()
root.wm_attributes('-topmost', 1)
file = filedialog.askopenfilename(
    title="{titulo}",
    filetypes=[("Excel", "*.xlsx *.xls"), ("Todos os arquivos", "*.*")]{ini_arg}
)
print(file)
'''
    return _run_tkinter_dialog(script)

# ──────────────────────────────────────────────────────────────────────────────
# PÁGINA
# ──────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    st.set_page_config(
        page_title="Auditoria TA · IA (VERIFIED FIXED 2026)",
        page_icon="🤖",
        layout="wide",
        initial_sidebar_state="expanded"
    )

# ──────────────────────────────────────────────────────────────────────────────
# CONSTANTES
# ──────────────────────────────────────────────────────────────────────────────
UNIDADES_AREAS = {
    "Juiz de Fora":    ["Utilidades","Aciaria","Laminação","Redução"],
    "Piracicaba":      ["Aciaria","Utilidades","Laminação"],
    "Sitrel":          ["Laminação"],
    "Mina do Andrade": ["Mineração"],
    "João Monlevade":  ["Aciaria","Laminação TL2","Laminação TL3","Utilidades","Alto Forno","Sinterização","GACAT"],
    "Trefilarias":     ["São Paulo","Sabará","Resende","Juiz de Fora"],
    "Sul Fluminense":  ["Aciaria (RS)","Utilidades (RS)","Laminação (RS)","Laminação (BM)","Utilidades (BM)"]
}

ESCALA = {0:("🔴","Não tem prática"),1:("🟠","Iniciando"),
          2:("🟡","Regular"),3:("🔵","Bom"),4:("🟢","Excelente")}

COR_NOTA = {0:"#D32F2F",1:"#E64A19",2:"#F9A825",3:"#1976D2",4:"#388E3C"}

# DB_PATH gerenciado pelo módulo database.py (C:\\AuditoriaTA\\dados\\)
DB_PATH = db_module.DB_PATH

# ──────────────────────────────────────────────────────────────────────────────
# CSS CUSTOMIZADO
# ──────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    st.markdown("""
<style>
/* Ocultar botão Deploy do Streamlit */
[data-testid="stDeployButton"] { display: none !important; }
.stDeployButton { display: none !important; }
button[kind="deployButton"] { display: none !important; }
[data-testid="stSidebar"] {background: #1a2744;}
[data-testid="stSidebar"] * {color: #e8eaf6 !important;}
[data-testid="stSidebar"] .stRadio label {color: #e8eaf6 !important;}
.card-subitem {
    background: #F8FAFF; border: 1px solid #D0DCF0;
    border-radius: 12px; padding: 16px 20px; margin: 10px 0;
    box-shadow: 0 2px 8px rgba(31,55,100,0.06);
}
.badge-ai {
    background: linear-gradient(90deg, #6C63FF, #3D5AFE);
    color: white; border-radius: 20px; padding: 4px 14px;
    font-size: 0.82em; font-weight: bold; display: inline-block;
}
.badge-nota {
    color: white; border-radius: 20px; padding: 5px 14px;
    font-weight: bold; display: inline-block; font-size: 1em;
}
.ai-result-box {
    background: #EEF2FF; border-left: 4px solid #3D5AFE;
    border-radius: 0 8px 8px 0; padding: 12px 16px; margin: 8px 0;
}
.check-item {color: #388E3C; margin: 2px 0;}
.miss-item  {color: #D32F2F; margin: 2px 0;}
.metric-card {
    background: white; border-radius: 10px; padding: 16px;
    text-align: center; box-shadow: 0 2px 8px rgba(0,0,0,0.08);
}
/* ── Anti-overlay: impede tela cinza durante carregamentos ── */
/* Remove o overlay semi-transparente que bloqueia a view durante st.spinner */
[data-testid="stSpinner"] > div:first-child {
    background: rgba(255,255,255,0.0) !important;
    position: relative !important;
    width: auto !important;
    height: auto !important;
}
[data-testid="stSpinner"] {
    background: transparent !important;
    box-shadow: none !important;
}
/* Mantém elementos visíveis durante execução de script */
.stApp > div {
    pointer-events: auto !important;
    opacity: 1 !important;
}
/* Spinner compacto sem bloquear toda a tela */
div.stSpinner {
    display: inline-flex !important;
    align-items: center;
    background: #f0f4ff !important;
    border: 1px solid #c5cef7 !important;
    border-radius: 8px !important;
    padding: 6px 12px !important;
    width: auto !important;
    max-width: 400px !important;
}
</style>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────────────────────
# DATABASE — delegado ao módulo database.py (persistência profissional)
# ──────────────────────────────────────────────────────────────────────────────
def init_db():
    """Inicializa / migra o banco. Faz backup automático na abertura."""
    db_module.init_db()
    # Backup automático a cada início de sessão (uma vez por session_state)
    if not st.session_state.get("_backup_feito"):
        bk = db_module.fazer_backup(motivo="sessao")
        if bk:
            st.session_state["_backup_feito"] = True

def get_db():
    return db_module.get_db()

# ── Helper centralizado: traduz erros de API OpenAI ─────────────────────────
def _mostrar_erro_openai(err, placeholder=None):
    """Detecta o tipo de erro OpenAI e exibe mensagem clara no Streamlit."""
    import streamlit as _st
    exibir = placeholder.error if placeholder else _st.error
    info   = placeholder.info  if placeholder else _st.info
    err_s  = str(err)

    # 401 – chave inválida / expirada
    if "401" in err_s or "Invalid or expired token" in err_s or \
       "invalid_api_key" in err_s or "Incorrect API key" in err_s or \
       "AuthenticationError" in err_s:
        exibir("❌ **Chave API inválida ou expirada.**")
        info(
            "**Como resolver:**\n"
            "1. Acesse [platform.openai.com/api-keys](https://platform.openai.com/api-keys)\n"
            "2. Clique em **+ Create new secret key**\n"
            "3. Copie a chave gerada\n"
            "4. Cole em ⚙️ Configurações → Configurar API Key → **Salvar Chave API**\n\n"
            "⚠️ Nunca compartilhe a chave em chats — o OpenAI revoga automaticamente chaves expostas."
        )
    # 429 – cota esgotada ou rate limit
    elif "insufficient_quota" in err_s or "429" in err_s or \
         "RateLimitError" in err_s:
        exibir("❌ **Cota da OpenAI esgotada ou limite de requisições atingido.**")
        info(
            "**Causas comuns:**\n"
            "- Orçamento do projeto OpenAI = $0 → acesse o projeto em "
            "platform.openai.com → Limits → defina Monthly budget\n"
            "- Saldo da conta zerado → Billing → Add to credit balance"
        )
    elif "model_not_found" in err_s or "does not have access" in err_s:
        exibir("❌ **Modelo não disponível para esta chave.**")
        info("Troque para Modo Econômico (GPT-4o-mini) em ⚙️ Configurações.")
    else:
        exibir(f"❌ Erro IA: {err_s[:200]}")

@st.cache_data(show_spinner=False, ttl=20)
def listar_auditorias():
    """Cache de 20s — evita N queries a cada rerun."""
    return db_module.listar_auditorias()

def criar_auditoria(unidade, area, ciclo, assessment_path, evidence_folder, api_key=""):
    return db_module.criar_auditoria(
        unidade, area, ciclo, assessment_path, evidence_folder, api_key)

def salvar_avaliacao(auditoria_id, pratica_num, pratica_nome, subitem_idx, subitem_nome,
                     evidencia_desc, n0,n1,n2,n3,n4, nota_sa,
                     decisao, nota_final, desc_nc, comentarios,
                     ia_decisao=None, ia_nota=None, ia_confianca=None,
                     ia_atendidos=None, ia_faltantes=None, ia_analise=None, ia_status=None):
    return db_module.salvar_avaliacao(
        auditoria_id, pratica_num, pratica_nome, subitem_idx, subitem_nome,
        evidencia_desc, n0,n1,n2,n3,n4, nota_sa,
        decisao, nota_final, desc_nc, comentarios,
        ia_decisao, ia_nota, ia_confianca,
        ia_atendidos, ia_faltantes, ia_analise, ia_status)

@st.cache_data(show_spinner=False, ttl=20)
def carregar_avaliacoes(auditoria_id):
    """Cache de 20s — evita query a cada rerun da página de auditoria."""
    return db_module.carregar_avaliacoes(auditoria_id)

@st.cache_data(show_spinner=False, ttl=20)
def get_auditoria(auditoria_id):
    """Cache de 20s — evita query a cada rerun/sidebar."""
    return db_module.get_auditoria(auditoria_id)

def atualizar_config(auditoria_id, assessment_path, evidence_folder, api_key):
    return db_module.atualizar_config(auditoria_id, assessment_path, evidence_folder, api_key)

# ── Invalidação de cache após escritas ─────────────────────────────────────
def _limpar_cache_auditoria():
    """Limpa todos os caches de dados após qualquer operação de escrita no banco.
    Deve ser chamada ANTES de st.rerun() quando dados foram modificados.
    """
    try:
        listar_auditorias.clear()
        get_auditoria.clear()
        carregar_avaliacoes.clear()
        construir_mapa_evidencias.clear()
    except Exception:
        pass  # seguro falhar silenciosamente

# ── PARSER PANDAS (Robust fallback for formulas) ──────────────────────────
def parse_assessment_pandas(file_path):
    import re
    import pandas as pd
    try:
        xl = pd.ExcelFile(file_path)
        sheet_name = next((n for n in xl.sheet_names if 'ROAD MAP' in n and 'Trefila' not in n), xl.sheet_names[0])
        df_sheet = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        
        praticas_dict = {}
        current_p_num = None
        current_s_idx = 0
        SKIP_KEYWORDS = {"EVIDÊNCIA", "SUBITEM", "DESCRIÇÃO", "PRÁTICA", "REQUISITO", "EVIDENCIAS", "N°", "NÃO TEM PRÁTICA"}

        def clean_num(v):
            if pd.isna(v): return ""
            s = str(v).strip()
            if s.endswith('.0'): s = s[:-2]
            return s

        for i, row in df_sheet.iterrows():
            row_vals = row.tolist()
            if all(pd.isna(v) for v in row_vals): continue
            
            first_col = clean_num(row_vals[0])
            col1_str = str(row_vals[1]).strip() if pd.notna(row_vals[1]) else ""
            col2_str = str(row_vals[2]).strip() if pd.notna(row_vals[2]) else ""
            
            # Detectar subitem ou prática
            sub_m = re.match(r'^(\d+)\.(\d+)', first_col)
            prac_m = re.match(r'^(\d+)\s*$', first_col)
            int_m  = re.match(r'^(\d+)\s*[-–]', first_col)
            prac_b = (not prac_m and not int_m and first_col.isdigit() and len(first_col) < 3 and col1_str)

            is_subitem = False
            if sub_m:
                p_num = int(sub_m.group(1)); s_idx = int(sub_m.group(2)) - 1
                current_p_num = p_num; current_s_idx = s_idx + 1; is_subitem = True
            elif (prac_m or int_m or prac_b) and len(first_col) < 50:
                m = prac_m or int_m
                p_num = int(m.group(1)) if m else int(first_col)
                p_nome = col1_str.replace('\n', ' ')
                current_p_num = p_num; current_s_idx = 0
                if p_num not in praticas_dict:
                    praticas_dict[p_num] = {'num': p_num, 'nome': p_nome or f"Prática {p_num}", 'subitems_map': {}}
                if col2_str and col2_str.upper() not in SKIP_KEYWORDS:
                    is_subitem = True; s_idx = current_s_idx; current_s_idx += 1
                else: continue
            elif current_p_num is not None and col2_str and not first_col:
                if col2_str.upper() not in SKIP_KEYWORDS and "N°" not in first_col:
                    p_num = current_p_num; s_idx = current_s_idx; current_s_idx += 1; is_subitem = True

            if is_subitem and current_p_num in praticas_dict:
                n_list = [str(row_vals[j]).strip() if pd.notna(row_vals[j]) else "" for j in range(3, 8)]
                final_nota = None
                for col_idx in [8, 9, 7]: # I, J, H
                    if len(row_vals) > col_idx:
                        v = row_vals[col_idx]
                        if pd.notna(v):
                            try: final_nota = int(float(v)); break
                            except: pass
                praticas_dict[current_p_num]['subitems_map'][s_idx] = {
                    'nome': col2_str.split('\n')[0].strip(),
                    'evidencia': col2_str,
                    'niveis': {k: v for k, v in enumerate(n_list)},
                    'nota_sa': final_nota
                }
        
        # Consolidar
        final_praticas = []
        for p_num in sorted(praticas_dict.keys()):
            p_data = praticas_dict[p_num]
            if not p_data['subitems_map']: continue
            max_idx = max(p_data['subitems_map'].keys())
            subitems = []
            for i in range(max_idx + 1):
                subitems.append(p_data['subitems_map'].get(i, {'nome': f'Subitem {i+1}', 'evidencia': '', 'niveis': {}, 'nota_sa': None}))
            final_praticas.append({'num': p_data['num'], 'nome': p_data['nome'], 'subitems': subitems})
        return final_praticas
    except Exception as e:
        print(f"Erro Pandas Parser: {e}")
        return []

# ──────────────────────────────────────────────────────────────────────────────
# PARSER DO ASSESSMENT
# ──────────────────────────────────────────────────────────────────────────────
def parse_assessment(file_path):
    import re
    try:
        wb = load_workbook(file_path, data_only=True)
        sheet_name = next((n for n in wb.sheetnames if 'ROAD MAP' in n and 'Trefila' not in n), wb.sheetnames[0])
        ws = wb[sheet_name]
        praticas_dict = {} # Use dict to store practices by number
        
        current_p_num = None
        current_s_idx = 0
        
        # Keywords to skip in Column C to avoid treating headers as subitems
        SKIP_KEYWORDS = {"EVIDÊNCIA", "SUBITEM", "DESCRIÇÃO", "PRÁTICA", "REQUISITO", "EVIDENCIAS", "N°", "NÃO TEM PRÁTICA"}

        # Diagnostic logging for Streamlit UI
        st.info(f"🔍 Analisando aba: {sheet_name}")
        pbar = st.progress(0)
        status_text = st.empty()

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            row_idx = i + 1
            if row_idx % 20 == 0:
                pbar.progress(min(row_idx / 500, 1.0))
                status_text.text(f"Processando linha {row_idx}...")
            
            if not row or all(v is None for v in row):
                continue
            
            # Robust first_col extraction
            raw_col_a = row[0]
            first_col = str(raw_col_a or "").strip()
            col1_str = str(row[1] or "").strip() # Column B
            col2_str = str(row[2] or "").strip() # Column C (Evidence/Description)
            
            # Practice detection
            subitem_match = re.match(r'^(\d+)\.(\d+)', first_col)
            practice_match = re.match(r'^(\d+)\s*$', first_col)
            integrated_practice_match = re.match(r'^(\d+)\s*[-–]', first_col)
            
            # Additional check: Practice might be in Column B if Column A is just the number
            practice_b_match = None
            if not practice_match and not integrated_practice_match:
                 if first_col.isdigit() and len(first_col) < 3 and col1_str:
                     practice_b_match = True

            if subitem_match:
                p_num = int(subitem_match.group(1))
                s_idx = int(subitem_match.group(2)) - 1
                current_p_num = p_num
                current_s_idx = s_idx + 1
                is_subitem = True
            elif (practice_match or integrated_practice_match or practice_b_match) and len(first_col) < 50:
                m = practice_match or integrated_practice_match
                p_num = int(m.group(1)) if m else int(first_col)
                p_nome = col1_str.replace('\n', ' ')
                current_p_num = p_num # Set the active practice
                current_s_idx = 0 
                
                if p_num not in praticas_dict:
                    praticas_dict[p_num] = {'num': p_num, 'nome': p_nome or f"Prática {p_num}", 'subitems_map': {}}
                elif p_nome and 'PRÁTICA' not in p_nome.upper():
                    praticas_dict[p_num]['nome'] = p_nome
                
                # Check if this row also contains a subitem (common in GACAT)
                if col2_str and col2_str.upper() not in SKIP_KEYWORDS:
                    is_subitem = True
                    p_num = current_p_num
                    s_idx = current_s_idx
                    current_s_idx += 1
                else:
                    st.write(f"✅ Prática {p_num} detectada: {p_nome[:40]}...")
                    continue # Skip to next row only if no subitem on this header row
            
            # Positional Subitem Detection (Smart Parser)
            elif current_p_num is not None and col2_str and not first_col:
                # Check if it's not a header row
                if col2_str.upper() not in SKIP_KEYWORDS and "N°" not in first_col:
                    p_num = current_p_num
                    s_idx = current_s_idx
                    current_s_idx += 1
                    is_subitem = True

            if is_subitem and p_num is not None and s_idx is not None:
                # Ensure practice exists
                if p_num not in praticas_dict:
                    praticas_dict[p_num] = {'num': p_num, 'nome': f"Prática {p_num}", 'subitems_map': {}}
                
                # Extract data
                n0, n1, n2, n3, n4 = row[3], row[4], row[5], row[6], row[7]
                
                # Robust note parsing
                final_nota = None
                for col_idx in [8, 9, 7]: # Columns I, J, H
                    if len(row) > col_idx:
                        val = db_module._safe_int(row[col_idx])
                        if val is not None:
                            final_nota = val
                            break
                
                praticas_dict[p_num]['subitems_map'][s_idx] = {
                    'nome': col2_str.split('\n')[0].strip(),
                    'evidencia': col2_str,
                    'niveis': {k: str(v).strip() if v else '' for k, v in enumerate([n0, n1, n2, n3, n4])},
                    'nota_sa': final_nota
                }

        # Convert dict to sorted list with ordered subitems
        final_praticas = []
        for p_num in sorted(praticas_dict.keys()):
            p_data = praticas_dict[p_num]
            subitems = []
            if not p_data['subitems_map']: continue
            
            max_idx = max(p_data['subitems_map'].keys())
            for i in range(max_idx + 1):
                subitems.append(p_data['subitems_map'].get(i, {'nome': f'Subitem {i+1}', 'evidencia': '', 'niveis': {}, 'nota_sa': None}))
            
            p_data['subitems'] = subitems
            del p_data['subitems_map']
            final_praticas.append(p_data)
            
        # TOTAL CHECK - If ZERO notes, try Pandas fallback
        total_notas = sum(1 for p in final_praticas for s in p['subitems'] if s['nota_sa'] is not None)
        if total_notas == 0:
            st.warning("⚠️ Openpyxl não encontrou notas (fórmulas?). Tentando motor Pandas...")
            p_fallback = parse_assessment_pandas(file_path)
            if p_fallback:
                total_f = sum(1 for p in p_fallback for s in p['subitems'] if s['nota_sa'] is not None)
                if total_f > 0:
                    st.success(f"✅ Motor Pandas recuperou {total_f} notas!")
                    return p_fallback
        
        st.success(f"📊 {len(final_praticas)} Práticas carregadas. Total notas SA: {total_notas}")
        return final_praticas
    except Exception as e:
        msg = f"Erro ao ler assessment: {e}"
        try: st.error(msg)
        except: print(msg)
        import traceback
        print(traceback.format_exc())
        return []

# ──────────────────────────────────────────────────────────────────────────────
# UTILITÁRIOS
# ──────────────────────────────────────────────────────────────────────────────
def calcular_nota_final(nota_sa, decisao, nota_livre=None):
    if nota_sa is None: return None
    if decisao=='permanece':    return nota_sa
    if decisao=='insuficiente':
        if nota_livre is not None:
            return max(0, min(int(nota_livre), max(0, nota_sa - 1)))
        return max(0, nota_sa - 1)
    if decisao=='inexistente':  return 0
    return None

def badge_nota_html(nota, extra=""):
    if nota is None: return '<span style="background:#aaa;color:white;padding:4px 12px;border-radius:20px;">?</span>'
    cor = COR_NOTA.get(int(nota),"#aaa")
    em,desc = ESCALA.get(int(nota),("",""))
    label = f"{em} {nota} — {desc}" + (f" {extra}" if extra else "")
    return f'<span class="badge-nota" style="background:{cor};">{label}</span>'

EXTS_IMG   = {'.jpg','.jpeg','.png','.gif','.bmp','.webp'}
EXTS_DOC   = {'.pdf','.xlsx','.xls','.docx','.doc'}
EXTS_VIDEO = {'.mp4','.avi','.mov','.mkv','.webm'}
EXTS_ALL   = EXTS_IMG | EXTS_DOC | EXTS_VIDEO

def listar_evidencias(folder):
    if not folder or not os.path.isdir(folder): return []
    return [f for f in sorted(Path(folder).rglob("*")) if f.is_file() and f.suffix.lower() in EXTS_ALL]

@st.cache_data(show_spinner=False)
def construir_mapa_evidencias(ev_folder: str) -> dict:
    """
    Varre a pasta de evidências UMA Única vez e retorna um dict:
      (pratica_num, sub_idx) -> [Path, ...]
    Cache automático pelo Streamlit enquanto ev_folder não mudar.
    """
    if not ev_folder or not os.path.isdir(ev_folder):
        return {}
    exts = EXTS_ALL
    mapa = {}

    for pasta_pratica in sorted(Path(ev_folder).iterdir()):
        if not pasta_pratica.is_dir(): continue
        # Detectar número da prática: "[1]", "1-", "1 ", "1."
        m_p = re.match(r'^\[?(\d+)[\]\-_\s\.]', pasta_pratica.name)
        if not m_p: continue
        p_num = int(m_p.group(1))

        for pasta_sub in sorted(pasta_pratica.iterdir()):
            if not pasta_sub.is_dir(): continue
            # Detectar número do sub-item: "1.1", "1.2", "2.1" ...
            m_s = re.match(r'^(\d+)\.(\d+)(?:[\s\-_\.]|$)', pasta_sub.name)
            if not m_s: continue
            s_num = int(m_s.group(2)) - 1  # converter para idx (0-based)
            arquivos = [
                f for f in sorted(pasta_sub.rglob("*"))
                if f.is_file() and f.suffix.lower() in exts
            ]
            mapa[(p_num, s_num)] = arquivos

    return mapa

def listar_evidencias_subitem(ev_folder, pratica_num, sub_idx):
    """Atalho que usa o mapa cacheado."""
    mapa = construir_mapa_evidencias(ev_folder)
    return mapa.get((pratica_num, sub_idx), [])

def exibir_evidencia(file_path):
    """Exibe o arquivo inline: imagem, PDF, DOCX, XLSX ou vídeo."""
    fp = Path(file_path)
    ext = fp.suffix.lower()
    if ext in EXTS_IMG:
        try:
            st.image(Image.open(fp), use_container_width=True)
        except Exception as e:
            st.error(f"Erro ao abrir imagem: {e}")
    elif ext == '.pdf':
        try:
            with open(fp, "rb") as f: b64 = base64.b64encode(f.read()).decode()
            st.markdown(
                f'<iframe src="data:application/pdf;base64,{b64}" '
                f'width="100%" height="700px" '
                f'style="border:1px solid #ddd;border-radius:8px;"></iframe>',
                unsafe_allow_html=True
            )
        except Exception as e:
            st.error(f"Erro ao abrir PDF: {e}")
    elif ext in {'.docx', '.doc'}:
        try:
            from docx import Document as DocxDocument
            doc = DocxDocument(fp)
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            if paragraphs:
                st.text_area(f"📄 {fp.name}", "\n\n".join(paragraphs), height=400)
            else:
                st.info("Documento sem texto extraível.")
        except Exception as e:
            st.warning(f"Não foi possível pré-visualizar: {e}")
            with open(fp, "rb") as f: data = f.read()
            st.download_button(f"⬇️ {fp.name}", data, fp.name)
    elif ext in {'.xlsx', '.xls'}:
        try:
            df_preview = pd.read_excel(fp, sheet_name=0)
            st.dataframe(df_preview, use_container_width=True, height=350)
        except Exception as e:
            st.warning(f"Não foi possível pré-visualizar: {e}")
            with open(fp, "rb") as f: data = f.read()
            st.download_button(f"⬇️ {fp.name}", data, fp.name)
    elif ext in EXTS_VIDEO:
        try:
            st.video(str(fp))
        except Exception as e:
            st.error(f"Erro ao reproduzir vídeo: {e}")
    else:
        with open(fp, "rb") as f: data = f.read()
        st.download_button(f"⬇️ {fp.name}", data, fp.name)

# ──────────────────────────────────────────────────────────────────────────────
# EXPORTADOR EXCEL
# ──────────────────────────────────────────────────────────────────────────────
def gerar_excel(auditoria_id):
    aud = get_auditoria(auditoria_id)
    df_aval = carregar_avaliacoes(auditoria_id)
    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = "Análise de Auditoria"

    # Cores para um visual mais profissional e de auditoria
    COR_TITULO_PRINCIPAL_FUNDO = "002060"  # Azul Escuro
    COR_TITULO_PRINCIPAL_FONTE = "FFFFFF"  # Branco
    COR_CABECALHO_FUNDO = "D9D9D9"       # Cinza Claro
    COR_CABECALHO_FONTE = "000000"       # Preto
    COR_PRATICA_FUNDO = "DDEBF7"         # Azul Médio Claro
    COR_PRATICA_FONTE = "002060"         # Azul Escuro
    COR_PADRAO_FONTE = "000000"          # Preto

    # Cores para formatação condicional
    COR_DIMINUI_FUNDO = "FF0000"         # Vermelho
    COR_AUMENTA_FUNDO = "00B0F0"         # Azul Claro (para diferenciar do azul escuro do título)
    COR_PERMANECE_FUNDO = "00B050"       # Verde Escuro
    COR_NAO_SE_APLICA_FUNDO = "FFC000"   # Laranja
    COR_EVID_INSUF_FUNDO = "FFFF00"      # Amarelo
    COR_EVID_INEXIST_FUNDO = "A6A6A6"    # Cinza Médio
    COR_FONTE_BRANCA = "FFFFFF"          # Branco

    def fill(h): return PatternFill(start_color=h, end_color=h, fill_type="solid")
    thin_side = Side(style='thin', color="000000")
    medium_side = Side(style='medium', color="000000")
    brd_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    brd_medium_left = Border(left=medium_side, right=thin_side, top=thin_side, bottom=thin_side)

    # 1. Título Principal (Linha 1)
    ws.merge_cells("A1:F1")
    ws["A1"].value = f"FORMULÁRIO PARA ANÁLISE DAS NOTAS DO SELF ASSESSMENT DA {aud['area'].upper()} DE {aud['unidade'].upper()}"
    ws["A1"].fill = fill(COR_TITULO_PRINCIPAL_FUNDO)
    ws["A1"].font = Font(name='Arial', size=14, bold=True, color=COR_TITULO_PRINCIPAL_FONTE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # 2. Cabeçalhos (Linha 2)
    headers = ["PRÁTICAS", "Nota", "Status da Nota", "Tipo de Não Conformidade", "Descrição da Não Conformidade", "Comentários Adicionais"]
    widths = [67, 17, 18, 28, 93, 107]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = fill(COR_CABECALHO_FUNDO)
        c.font = Font(name='Arial', size=10, bold=True, color=COR_CABECALHO_FONTE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = brd_thin
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 25

    # 3. Conteúdo
    pratica_atual = None; row = 3
    status_map = {"permanece": "Permanece", "insuficiente": "Diminui", "inexistente": "Diminui", "aumenta": "Aumenta", "pendente": ""}
    tipo_map = {'permanece': 'Não se aplica', 'insuficiente': 'Evidências insuficiente', 'inexistente': 'Evidências inexistente', 'pendente': ''}

    for _, av in df_aval.iterrows():
        # Linha de Prática (Título da Seção)
        if av['pratica_num'] != pratica_atual:
            pratica_atual = av['pratica_num']
            ws.merge_cells(f"A{row}:F{row}")
            c = ws[f"A{row}"]
            c.value = f"{av['pratica_num']} - {av['pratica_nome'].upper()}"
            c.fill = fill(COR_PRATICA_FUNDO)
            c.font = Font(name='Arial', size=10, bold=True, color=COR_PRATICA_FONTE)
            c.alignment = Alignment(vertical="center")
            c.border = brd_thin
            ws.row_dimensions[row].height = 20
            row += 1

        # Linha de Subitem
        d = av['decisao'] or 'pendente'
        
        # Valores das colunas
        # Coluna B = Nota do Self Assessment (cópia da coluna I da planilha Assessment)
        vals = [
            av['subitem_nome'],
            av['nota_self_assessment'] if pd.notna(av['nota_self_assessment']) else '',
            status_map.get(d, ''),
            tipo_map.get(d, ''),
            av.get('descricao_nc', '') or '',
            av.get('comentarios', '') or ''
        ]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = Font(name='Arial', size=10, color=COR_PADRAO_FONTE)
            c.alignment = Alignment(wrap_text=True, vertical="center")
            c.border = brd_medium_left if ci == 1 else brd_thin

            # Formatação condicional para Status da Nota (Coluna C, ci=3)
            if ci == 3:
                if val == "Diminui":
                    c.fill = fill(COR_DIMINUI_FUNDO)
                    c.font = Font(name='Arial', size=10, bold=True, color=COR_FONTE_BRANCA)
                elif val == "Aumenta": # Adicionado para o caso de 'Aumenta'
                    c.fill = fill(COR_AUMENTA_FUNDO)
                    c.font = Font(name='Arial', size=10, bold=True, color=COR_FONTE_BRANCA)
                elif val == "Permanece":
                    c.fill = fill(COR_PERMANECE_FUNDO)
                    c.font = Font(name='Arial', size=10, bold=True, color=COR_FONTE_BRANCA)
                c.alignment = Alignment(horizontal="center", vertical="center")

            # Formatação condicional para Tipo de Não Conformidade (Coluna D, ci=4)
            elif ci == 4:
                if val == "Não se aplica":
                    c.fill = fill(COR_NAO_SE_APLICA_FUNDO)
                elif val == "Evidências insuficiente":
                    c.fill = fill(COR_EVID_INSUF_FUNDO)
                elif val == "Evidências inexistente":
                    c.fill = fill(COR_EVID_INEXIST_FUNDO)
                c.alignment = Alignment(horizontal="center", vertical="center")
            
            # Centralizar Nota (Coluna B, ci=2)
            elif ci == 2:
                c.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajustar altura da linha com base no conteúdo (estimativa simples)
        desc_len = len(str(vals[4])) if vals[4] else 0
        coment_len = len(str(vals[5])) if vals[5] else 0
        max_len = max(desc_len, coment_len)
        num_linhas = max(1, max_len / 80)
        ws.row_dimensions[row].height = max(30, num_linhas * 15)
        row += 1

    out = io.BytesIO(); wb.save(out); out.seek(0); return out

# ──────────────────────────────────────────────────────────────────────────────
# INICIALIZAÇÃO
# ──────────────────────────────────────────────────────────────────────────────
init_db()

for key,val in [('pagina','dashboard'),('auditoria_id',None),('praticas_cache',{})]:
    if key not in st.session_state: st.session_state[key]=val

# ──────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ──────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style='text-align:center;padding:12px 0 6px 0;'>
        <div style='font-size:2.5em;'>🤖</div>
        <div style='font-size:1.1em;font-weight:bold;color:#90CAF9;'>Auditoria TA · IA</div>
        <div style='font-size:0.75em;color:#78909C;margin-top:3px;'>Automação Industrial Longos</div>
    </div>
    <hr style='border-color:#3D5068;margin:10px 0;'>
    """, unsafe_allow_html=True)

    menu = st.radio("nav", ["🏠 Dashboard","➕ Nova Auditoria","📋 Auditar",
                              "📊 Relatórios","📔 Diário de Auditoria",
                              "📁 Histórico de Auditorias",
                              "🗄️ Dados & Histórico","⚙️ Configurações"],
                    label_visibility="collapsed")
    pages = {"🏠 Dashboard":"dashboard","➕ Nova Auditoria":"nova",
             "📋 Auditar":"auditar","📊 Relatórios":"relatorios",
             "📔 Diário de Auditoria":"diario",
             "📁 Histórico de Auditorias":"historico",
             "🗄️ Dados & Histórico":"dados","⚙️ Configurações":"config"}
    st.session_state.pagina = pages[menu]

    if st.session_state.auditoria_id:
        aud = get_auditoria(st.session_state.auditoria_id)
        if aud:
            st.markdown("---")
            st.markdown(f"""
            <div style='background:#243455;border-radius:8px;padding:10px;font-size:0.82em;'>
                <b style='color:#90CAF9;'>Auditoria Ativa:</b><br>
                📍 {aud['unidade']}<br>
                🏭 {aud['area']}<br>
                🔄 Ciclo {aud.get('ciclo','')}
            </div>""", unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────────────────────
# DASHBOARD
# ──────────────────────────────────────────────────────────────────────────────
@st.fragment
def _render_subitem_frag(aud_id: int, pratica_num: int, pratica_nome: str,
                         sub_idx: int, api_key: str, aud: dict):
    """Renderiza conteúdo de um sub-item como fragment — apenas este bloco
    reroda ao salvar decisão/chat/consenso, sem rerun de página inteira."""
    # ── Re-buscar dados frescos a cada rerun do fragment ──────────────
    df_frag = carregar_avaliacoes(aud_id)
    av_rows = df_frag[
        (df_frag['pratica_num'] == pratica_num) &
        (df_frag['subitem_idx'] == sub_idx)
    ]
    if av_rows.empty:
        st.caption("⚠️ Sub-item sem dados.")
        return
    av = av_rows.iloc[0]

    sub_nome       = av['subitem_nome']
    nota_sa        = av['nota_self_assessment']
    decisao_atual  = av['decisao'] or 'pendente'
    desc_nc_atual  = av.get('descricao_nc', '') or ''
    coment_atual   = av.get('comentarios', '') or ''
    ev_desc        = av.get('evidencia_descricao', '') or ''
    ia_status_val  = av.get('ia_status')
    kp             = f"p{pratica_num}_s{sub_idx}"

    # Re-buscar mapa de evidências (cached — custo ≈ 0)
    ev_folder  = aud.get('evidence_folder_path', '') or ''
    ev_mapa_cache = construir_mapa_evidencias(ev_folder)

    # Re-buscar chat deste sub-item (1 query direta — rápida)
    chat_msgs = db_module.carregar_chat(aud_id, pratica_num, sub_idx)

    # ── Pre-buscar dados de critérios (PO.AUT.002) ──────────────
    try:
        from criterios_oficiais import get_criterio as _gc_crit, REGRAS_GERAIS as _RG_crit
        _crit_data = _gc_crit(pratica_num, sub_idx)
        _niveis_main = _crit_data.get('niveis', {})
        _ev_main     = _crit_data.get('evidencias_exigidas', '') or ''
        _re_main     = _crit_data.get('regras_especiais', '') or ''
    except Exception:
        _niveis_main = {}; _ev_main = ''; _re_main = ''; _RG_crit = ''

    st.session_state[f"{kp}_open"] = True  # mantém aberto nas reruns enquanto visível
    # Âncora HTML para auto-scroll
    st.markdown(f'<div id="anchor_{kp}"></div>', unsafe_allow_html=True)
    col_esq, col_dir = st.columns([1,1])

    # ────────── COLUNA ESQUERDA: EVIDÊNCIAS ──────────
    with col_esq:
        st.markdown("##### 📁 Evidências")
        # Mapa já foi construído UMA vez antes do loop de práticas
        arq_subitem = ev_mapa_cache.get((pratica_num, sub_idx), [])

        if arq_subitem:
            imgs   = [f for f in arq_subitem if f.suffix.lower() in {'.jpg','.jpeg','.png','.bmp','.webp','.gif'}]
            outros = [f for f in arq_subitem if f not in imgs]

            # Contagem sempre visível (sem renderizar imagens)
            partes = []
            if imgs:   partes.append(f"📷 {len(imgs)} imagem(ns)")
            if outros: partes.append(f"📄 {len(outros)} documento(s)")
            st.markdown(" — ".join(partes))

            # ── GALERIA DE IMAGENS (lightbox HTML com clique direto) ──
            if imgs:
                with st.expander(f"📷 Galeria de imagens ({len(imgs)}) — clique para ampliar", expanded=True):
                    render_image_gallery_html(imgs, kp)

            # ── DOCUMENTOS E VÍDEOS (preview inline) ──
            videos = [f for f in arq_subitem if f.suffix.lower() in EXTS_VIDEO]
            docs   = [f for f in arq_subitem if f.suffix.lower() in EXTS_DOC]
            if videos:
                with st.expander(f"🎬 Vídeos ({len(videos)})", expanded=False):
                    for vf in videos:
                        st.caption(f"▶ {vf.name}")
                        try:
                            st.video(str(vf))
                        except Exception as ex:
                            st.warning(f"Não foi possível reproduzir: {ex}")
            if docs:
                with st.expander(f"📄 Documentos ({len(docs)})", expanded=False):
                    sel_doc = st.selectbox(
                        "Selecione o documento para pré-visualizar:",
                        ["(selecione)"] + [f.name for f in docs],
                        key=f"{kp}_sel_doc"
                    )
                    if sel_doc != "(selecione)":
                        fp_doc = next((f for f in docs if f.name == sel_doc), None)
                        if fp_doc:
                            exibir_evidencia(fp_doc)
                            with open(fp_doc, "rb") as _fh:
                                st.download_button(
                                    f"⬇️ Baixar {fp_doc.name}",
                                    _fh.read(), fp_doc.name,
                                    key=f"{kp}_dl_{fp_doc.name[:20]}"
                                )

        elif ev_folder:
            st.caption(f"📭 Pasta {pratica_num}.{sub_idx+1} não encontrada ou vazia")
            st.caption(f"📂 Pasta de evidências: {ev_folder}")
        else:
            st.info("⚙️ Configure a pasta de evidências em ⚙️ Configurações.")

        # ── Análise IA individual ──
        if api_key:
            if st.button(f"🤖 Analisar este subitem com IA", key=f"{kp}_ia"):
                try:
                    from ai_analyzer import AuditAIAnalyzer
                    # Usa ev_mapa_cache (mesmos arquivos exibidos na UI)
                    files=ev_mapa_cache.get((pratica_num,sub_idx),[])
                    _eco = st.session_state.get('modo_ia','completo') == 'economico'
                    analyzer=AuditAIAnalyzer(api_key, economico=_eco)
                    # Feedback antecipado: quantos arquivos serão analisados
                    _n_imgs_prev = sum(1 for f in files if f.suffix.lower() in {'.jpg','.jpeg','.png','.gif','.webp','.bmp'})
                    _n_docs_prev = sum(1 for f in files if f.suffix.lower() in {'.pdf','.docx','.doc','.xlsx','.xls'})
                    _spin_msg = f"🤖 Analisando {_n_docs_prev} documento(s) + {_n_imgs_prev} imagem(ns) com GPT-4o..."
                    if not files:
                        _spin_msg = "🤖 Analisando (sem arquivos de evidência encontrados)..."
                    _ia_status_ph = st.empty()
                    _ia_status_ph.info(_spin_msg)
                    res=analyzer.analyze_subitem(
                        pratica_num=pratica_num,
                        subitem_idx=sub_idx,
                        pratica_nome=pratica_nome,
                        subitem_nome=sub_nome,
                        evidencia_descricao=ev_desc,
                        niveis_planilha={k:av.get(f'nivel_{k}','') or '' for k in range(5)},
                        nota_self_assessment=nota_sa or 0,
                        evidence_files=files)
                    _ia_status_ph.empty()
                    ia_dec=res.get('decisao','pendente')
                    nota_final=calcular_nota_final(nota_sa,ia_dec)
                    salvar_avaliacao(
                        aud_id,pratica_num,pratica_nome,sub_idx,sub_nome,ev_desc,
                        av.get('nivel_0',''),av.get('nivel_1',''),av.get('nivel_2',''),
                        av.get('nivel_3',''),av.get('nivel_4',''),nota_sa,
                        ia_dec,nota_final,res.get('descricao_nc',''),res.get('comentarios',''),
                        ia_dec,res.get('nota_sugerida'),res.get('confianca'),
                        res.get('pontos_atendidos',[]),res.get('pontos_faltantes',[]),
                        res.get('analise_detalhada',''),'ok')
                    # Exibir cobertura da análise (sem rerun imediato, mostrar 1s)
                    _cob = res.get('cobertura_relatorio','')
                    if _cob:
                        st.caption(_cob)
                    _limpar_cache_auditoria()
                    st.toast("✅ Análise concluída!", icon="✅"); st.rerun()
                except ImportError as e:
                    st.error("❌ Dependência não instalada. Execute: `pip install openai pdfplumber python-docx`")
                except Exception as e:
                    err_str = str(e)
                    _mostrar_erro_openai(e)

    # ────────── COLUNA DIREITA: AVALIAÇÃO ──────────
    with col_dir:
        st.markdown("##### 📊 Avaliação")

        # Nota SA
        if nota_sa is not None:
            em,ed=ESCALA.get(int(nota_sa),("",""))
            st.markdown(f"**Nota Self Assessment:** {em} `{nota_sa}` — {ed}")

        # ── CRITÉRIOS E REGRAS (PO.AUT.002) ──────────────────────────────────────────
        try:
            # 1. Regras Especiais (Destaque conforme sistema antigo - Amarelo/Aviso)
            if _re_main:
                st.markdown(f"""
                <div style="background-color:#FFF3E0; border-left:5px solid #FF9800; padding:10px; border-radius:4px; margin-bottom:15px;">
                    <span style="color:#E65100; font-weight:bold;">⚠️ Regras Especiais:</span><br>
                    <span style="color:#4E342E;">{_re_main}</span>
                </div>
                """, unsafe_allow_html=True)
            
            # 2. Critérios (Expandido por padrão para facilitar a consulta)
            if _niveis_main or _ev_main:
                with st.expander("📖 Critérios de Avaliação (PO.AUT.002)", expanded=True):
                    # ── Sub-tabs: Níveis | Evidência ──
                    _tab_nv, _tab_ev = st.tabs(["📊 Níveis 0–4", "📋 Evidência Exigida"])

                    with _tab_nv:
                        _nomes_nv_main = {
                            0: "🔴 Nível 0",
                            1: "🟠 Nível 1",
                            2: "🟡 Nível 2",
                            3: "🔵 Nível 3",
                            4: "🟢 Nível 4",
                        }
                        _rows_nv_main = []
                        for _nv_m in range(5):
                            _t_m = _niveis_main.get(_nv_m, av.get(f'nivel_{_nv_m}', '') or '')
                            if _t_m:
                                _rows_nv_main.append({"Nível": _nomes_nv_main[_nv_m], "Descrição": _t_m})
                        
                        if _rows_nv_main:
                            import pandas as _pd_nv
                            # st.table para visual estático e limpo
                            st.table(_pd_nv.DataFrame(_rows_nv_main))

                    with _tab_ev:
                        if _ev_main:
                            _ev_itens = []
                            for _ep in _ev_main.split(';'):
                                for _eq in _ep.splitlines():
                                    _eq = _eq.strip(' -\t')
                                    if _eq: _ev_itens.append(_eq)
                            if not _ev_itens: _ev_itens = [_ev_main.strip()]
                            
                            import pandas as _pd_ev
                            _df_ev = _pd_ev.DataFrame(
                                [{"#": i+1, "Item de Evidência": txt} for i, txt in enumerate(_ev_itens)]
                            )
                            st.table(_df_ev)
                        else:
                            st.caption("Nenhuma evidência específica registrada.")

            # 3. Regras Gerais (Abaixo dos critérios, recolhido por padrão)
            with st.expander("🖍️ Regras Gerais (Todas as Práticas)", expanded=False):
                st.markdown(_RG_crit)

        except Exception as e_crit:
            st.error(f"Erro ao carregar critérios: {e_crit}")

        # ── CHECKLIST DE CONFERÊNCIA (PO.AUT.002 – literal do documento) ────
        try:
            from checklist_po_aut_002 import get_checklist as _gck
            _ck      = _gck(pratica_num, sub_idx)
            _ck_ver  = _ck.get('verificar', [])
            _ck_arm  = _ck.get('armadilhas', [])
            _ck_nv4  = _ck.get('nota4', '')
            _ck_reg  = _ck.get('regras', [])
            _ck_hard = _ck.get('hard_rule')
            if _ck_ver or _ck_arm:
                with st.expander("📋 Checklist de Conferência (PO.AUT.002)", expanded=False):

                    # ── 1. REGRA HARD (alerta vermelho no topo) ──
                    if _ck_hard:
                        st.error(f"🚨 **REGRA RÍGIDA — NOTA 0 AUTOMÁTICA:** {_ck_hard}")

                    # ── 2. O QUE VERIFICAR (checkboxes principais) ──
                    if _ck_ver:
                        st.markdown("**🔍 O que verificar:**")
                        for _i_v, _item_v in enumerate(_ck_ver):
                            st.checkbox(_item_v, key=f"{kp}_ck_v_{_i_v}")

                    # ── 3. REGRAS ESPECIAIS ──
                    _reg_filtradas = [r for r in _ck_reg if r not in ('--', '-')]
                    if _reg_filtradas:
                        st.markdown("**⚠️ Regras Especiais:**")
                        for _r in _reg_filtradas:
                            st.warning(_r)

                    # ── 4. CRITÉRIO NOTA 4 ──
                    if _ck_nv4:
                        st.info(f"🟢 **Critério Nota 4:** {_ck_nv4}")

                    # ── 5. ARMADILHAS (checkboxes de atenção) ──
                    if _ck_arm:
                        st.markdown("**🪤 Armadilhas:**")
                        for _i_a, _item_a in enumerate(_ck_arm):
                            st.checkbox(_item_a, key=f"{kp}_ck_a_{_i_a}")

        except Exception:
            pass  # Silencia erros para não quebrar a UI

        # Resultado IA (se disponível)
        if ia_status_val=='ok':
            ia_dec_val=av.get('ia_decisao','')
            ia_nota_val=av.get('ia_nota_sugerida')
            ia_conf=av.get('ia_confianca','')
            ia_analise=av.get('ia_analise_detalhada','') or ''
            try: ia_atend=json.loads(av.get('ia_pontos_atendidos','[]') or '[]')
            except: ia_atend=[]
            try: ia_falt=json.loads(av.get('ia_pontos_faltantes','[]') or '[]')
            except: ia_falt=[]

            conf_color={'alta':'#388E3C','media':'#F57C00','baixa':'#D32F2F'}.get(ia_conf,'#666')
            decisao_label={'permanece':'✅ Permanece','insuficiente':'⚠️ Insuficiente','inexistente':'❌ Inexistente'}.get(ia_dec_val, ia_dec_val)
            st.markdown(f"""
            <div class="ai-result-box">
                <b>🤖 Parecer da IA (Critérios PO.AUT.002)</b> &nbsp;
                <span style="color:{conf_color};font-size:0.85em;">&#x25cf; Confiança: {ia_conf.upper()}</span><br>
                <b>Decisão sugerida:</b> {decisao_label} &nbsp;|&nbsp;
                <b>Nota sugerida:</b> {ia_nota_val}
            </div>""", unsafe_allow_html=True)

            if ia_atend or ia_falt:
                with st.expander("📋 Pontos analisados pela IA", expanded=True):
                    if ia_atend:
                        st.markdown("**✅ Atendido:**")
                        for p in ia_atend:
                            st.markdown(f'<div class="check-item">✔ {p}</div>',unsafe_allow_html=True)
                    if ia_falt:
                        st.markdown("**❌ Faltante:**")
                        for p in ia_falt:
                            st.markdown(f'<div class="miss-item">✖ {p}</div>',unsafe_allow_html=True)

            if ia_analise:
                with st.expander("🔍 Análise detalhada da IA"):
                    # ── Inventário de imagens (tabela) ──
                    if "━━━ INVENTÁRIO DE IMAGENS" in ia_analise:
                        _partes = ia_analise.split("\n\n━━━ INVENTÁRIO DE IMAGENS")
                        _texto_analise = _partes[0].strip()
                        _inv_raw = _partes[1] if len(_partes) > 1 else ""
                        if _texto_analise:
                            st.write(_texto_analise)
                        if _inv_raw:
                            st.markdown("---")
                            st.markdown("**📷 Inventário de Imagens Analisadas:**")
                            _inv_linhas = [l for l in _inv_raw.strip().splitlines() if l.strip() and not l.startswith("━")]
                            _inv_data = []
                            for _il in _inv_linhas:
                                _icon = "✅" if _il.startswith("✅") else "⚠️" if _il.startswith("⚠") else "📸"
                                _il_clean = _il.lstrip("✅⚠️📸 ")
                                _parts_il = _il_clean.split(" → ", 1)
                                _arq = _parts_il[0].strip() if _parts_il else _il_clean
                                _eq  = _parts_il[1].strip() if len(_parts_il) > 1 else "—"
                                _eq_parts = _eq.split(" | ", 1)
                                _equip = _eq_parts[0].strip()
                                _obs   = _eq_parts[1].strip() if len(_eq_parts) > 1 else ""
                                _inv_data.append({"Status": _icon, "Arquivo": _arq, "Equipamento": _equip, "Observação": _obs})
                            if _inv_data:
                                import pandas as _pd_inv
                                _df_inv = _pd_inv.DataFrame(_inv_data)
                                st.dataframe(
                                    _df_inv, use_container_width=True, hide_index=True,
                                    column_config={
                                        "Status":      st.column_config.TextColumn(width="small"),
                                        "Arquivo":     st.column_config.TextColumn(width="medium"),
                                        "Equipamento": st.column_config.TextColumn(width="medium"),
                                        "Observação":  st.column_config.TextColumn(width="large"),
                                    }
                                )
                    else:
                        st.write(ia_analise)

            # ── BANNER DE DIVERGÊNCIA AUTOMÁTICO ──
            _nota_ia_num = av.get('ia_nota_sugerida')
            _nota_sa_num = nota_sa
            try:
                _nota_ia_num = int(float(_nota_ia_num)) if _nota_ia_num is not None else None
            except (ValueError, TypeError):
                _nota_ia_num = None
            _pts_faltantes_raw = av.get('ia_pontos_faltantes','[]') or '[]'
            try:
                _pts_faltantes = json.loads(_pts_faltantes_raw) if isinstance(_pts_faltantes_raw, str) else _pts_faltantes_raw
            except Exception:
                _pts_faltantes = []

            _divergencia = False
            _motivo_div  = ""
            if _nota_ia_num is not None and _nota_sa_num is not None:
                _diff = _nota_sa_num - _nota_ia_num
                if _diff >= 2:
                    _divergencia = True
                    _motivo_div = f"IA sugeriu nota **{_nota_ia_num}**, mas SA declarou **{_nota_sa_num}** (diferença de {_diff} pontos)."
                elif _pts_faltantes and _nota_ia_num >= 3:
                    _divergencia = True
                    _motivo_div = f"IA atribuiu nota **{_nota_ia_num}** mas identificou **{len(_pts_faltantes)} ponto(s) faltante(s)**. Evidência pode ser insuficiente."

            if _divergencia:
                st.warning(f"⚠️ **Divergência detectada:** {_motivo_div}  \nRevise colaborativamente com a IA para justificar a decisão final.")

            st.info("💡 Aceite a sugestão da IA, sobrescreva manualmente abaixo, ou **revise colaborativamente** com a IA.")

            # ═══════════════════════════════════════════════════════
            # PAINEL DE REVISÃO COLABORATIVA HUMANO-IA
            # ═══════════════════════════════════════════════════════
            chat_key = f"{kp}_chat_aberto"
            if chat_key not in st.session_state:
                st.session_state[chat_key] = False

            # chat_msgs já foi carregado no topo da função via db_module.carregar_chat
            n_msgs = len([m for m in chat_msgs if m["role"] != "sistema"])
            btn_label = f"💬 Revisar com IA ({n_msgs} mensagens)" if n_msgs > 0 else "💬 Revisar com IA — Dialogar sobre divergência"

            col_chat_btn, col_chat_clr = st.columns([3, 1])
            with col_chat_btn:
                if st.button(btn_label, key=f"{kp}_open_chat", use_container_width=True):
                    st.session_state[chat_key] = not st.session_state[chat_key]
            with col_chat_clr:
                if n_msgs > 0:
                    if st.button("🗑️ Limpar chat", key=f"{kp}_clear_chat", use_container_width=True):
                        db_module.limpar_chat(aud_id, pratica_num, sub_idx)
                        _limpar_cache_auditoria()
                        st.session_state[f'prat_{pratica_num}_open'] = True
                        st.session_state[f'{kp}_open'] = True
                        st.session_state[chat_key] = True
                        st.rerun(scope="fragment")

            _rev_exp_open = st.session_state.get(chat_key, False)
            with st.expander("💬 Revisão Colaborativa Humano–IA", expanded=_rev_exp_open):
                    st.caption("Apresente suas observações de campo. A IA revisa sua análise considerando o contexto adicional.")

                    # ── Exibir histórico do chat ──
                    if chat_msgs:
                        for msg in chat_msgs:
                            role = msg["role"]
                            if role == "sistema":
                                st.caption(f"⚙️ {msg.get('conteudo', msg.get('content', ''))}")
                                continue
                            ts = msg.get("timestamp","")
                            try:
                                from datetime import datetime as _dt
                                ts_fmt = _dt.fromisoformat(ts).strftime("%d/%m %H:%M")
                            except Exception:
                                ts_fmt = ts[:16]

                            if role == "user":
                                st.markdown(f"""
    <div style="background:#E3F2FD;border-radius:12px 12px 4px 12px;
padding:10px 14px;margin:6px 0 6px 40px;
border-left:3px solid #1565C0;">
    <b>👤 Auditor</b> <span style="color:#888;font-size:11px">{ts_fmt}</span><br>
    {msg.get('conteudo', msg.get('content', ''))}
    </div>""", unsafe_allow_html=True)
                            else:
                                dec_snap = msg.get("decisao_snapshot")
                                nota_snap = msg.get("nota_snapshot")
                                snap_info = ""
                                if dec_snap:
                                    snap_icon = {"permanece":"✅","insuficiente":"⚠️","inexistente":"❌"}.get(dec_snap,"")
                                    snap_info = f'<br><span style="font-size:11px;color:#555;">→ Posição revisada: {snap_icon} {dec_snap} | Nota: {nota_snap}</span>'
                                st.markdown(f"""
    <div style="background:#F3E5F5;border-radius:12px 12px 12px 4px;
padding:10px 14px;margin:6px 40px 6px 0;
border-left:3px solid #6A1B9A;">
    <b>🤖 IA Auditora</b> <span style="color:#888;font-size:11px">{ts_fmt}</span><br>
    {msg.get('conteudo', msg.get('content', ''))}{snap_info}
    </div>""", unsafe_allow_html=True)
                    else:
                        st.caption("Nenhuma mensagem ainda. Inicie a revisão abaixo.")

                    st.markdown("---")
                    # ── Input de nova mensagem ──
                    sugestoes = [
                        "(escreva sua observação...)",
                        "Há jobs com erro no sistema de backup — os backups não estão sendo executados.",
                        "Os supervisórios (SCADA) não possuem backup configurado, apenas os PLCs.",
                        "As evidências mostram apenas nomes de arquivos, sem comprovação da rotina.",
                        "A data das evidências está desatualizada, não reflete o período auditado.",
                        "Confirmo o parecer da IA. A evidência está correta.",
                    ]
                    nova_msg_input = st.text_area(
                        "✍️ Sua observação para a IA:",
                        height=100,
                        key=f"{kp}_chat_input",
                        placeholder="Ex: Há jobs com erro. Os backups dos supervisórios não existem. A evidência é insuficiente porque..."
                    )

                    # Sugestões rápidas
                    st.markdown("**Sugestões rápidas:**")
                    sug_cols = st.columns(2)
                    for si, sug in enumerate(sugestoes[1:], 0):
                        with sug_cols[si % 2]:
                            if st.button(f"💡 {sug[:60]}…" if len(sug)>60 else f"💡 {sug}",
                                         key=f"{kp}_sug_{si}",
                                         use_container_width=True):
                                # Pré-preencher via session_state — rerun mínimo sem limpar cache
                                st.session_state[f"{kp}_chat_input"] = sug
                                st.session_state[f'prat_{pratica_num}_open'] = True
                                st.session_state[f'{kp}_open'] = True
                                st.session_state[chat_key] = True
                                st.rerun(scope="fragment")  # necessário para atualizar o text_area visualmente

                    # (pop de _chat_input_val removido — widget key gerencia o valor)

                    col_send, col_apply = st.columns([1, 1])
                    # Botão "Aplicar consenso" — usa última sugestão da IA no chat
                    with col_apply:
                        _ultima_ia = next((m for m in reversed(chat_msgs) if m["role"]=="assistant"), None)
                        _dec_cons   = _ultima_ia.get("decisao_snapshot") if _ultima_ia else None
                        _nota_cons  = _ultima_ia.get("nota_snapshot")    if _ultima_ia else None
                        # ── Resolver conflito nota ↔ decisão (a nota vence) ──
                        if _nota_cons is not None:
                            _nc_int = int(_nota_cons) if str(_nota_cons).lstrip('-').isdigit() else None
                            if _nc_int is not None:
                                if _nc_int == 0 and _dec_cons != "inexistente":
                                    _dec_cons = "inexistente"
                                elif _nc_int > 0 and _nc_int < (nota_sa or 5) and _dec_cons == "permanece":
                                    _dec_cons = "insuficiente"
                        _apply_disabled = (_dec_cons is None)
                        _apply_label = (
                            f"✅ Aplicar consenso ({_dec_cons} | Nota {_nota_cons})"
                            if _dec_cons else "✅ Aplicar consenso"
                        )
                        if st.button(_apply_label, key=f"{kp}_apply_cons",
                                     use_container_width=True, disabled=_apply_disabled):
                            _nota_ap = _nota_cons if _nota_cons is not None else calcular_nota_final(nota_sa, _dec_cons)
                            # ── Observação do auditor (mensagens user no chat) ──
                            _obs_ap = " | ".join(
                                m.get("conteudo", m.get("content", ""))[:200] for m in chat_msgs if m["role"]=="user"
                            )
                            # ── Resposta/justificativa da IA (última msg assistant) ──
                            _resp_ia_cons = _ultima_ia.get("conteudo", _ultima_ia.get("content", "")) if _ultima_ia else ""
                            # Limpar marcadores técnicos internos da resposta
                            import re as _re_cons
                            _resp_limpa = _re_cons.sub(
                                r'REVIS[ÃA]O_[A-Z_ÇÃÁÉÍÓÚ]+:\s*\S+', "", _resp_ia_cons
                            ).strip()
                            # ── Montar desc_nc: sempre salva justificativa IA (UI exibe só quando insuficiente/inexistente) ──
                            _desc_nc_cons = _resp_limpa[:1500] if _resp_limpa else _obs_ap[:500]
                            # ── Comentário automático baseado na decisão do consenso ──
                            if _dec_cons == "permanece":
                                _coment_cons = "Nota permanece"
                            elif _dec_cons == "inexistente":
                                _coment_cons = "Nota passa para 0"
                            elif _dec_cons == "insuficiente":
                                _coment_cons = f"Nota passa para {_nota_ap}" if _nota_ap is not None else "Evidência insuficiente"
                            else:
                                _coment_cons = (
                                    f"Revisão colaborativa: {_obs_ap[:300]}"
                                    if _obs_ap else av.get("comentarios", "")
                                )
                            salvar_avaliacao(
                                aud_id, pratica_num, pratica_nome,
                                sub_idx, sub_nome, ev_desc,
                                av.get("nivel_0",""), av.get("nivel_1",""),
                                av.get("nivel_2",""), av.get("nivel_3",""),
                                av.get("nivel_4",""), nota_sa,
                                _dec_cons, _nota_ap,
                                _desc_nc_cons,
                                _coment_cons,
                                av.get("ia_decisao"), av.get("ia_nota_sugerida"),
                                av.get("ia_confianca"),
                                json.loads(av.get("ia_pontos_atendidos","[]") or "[]"),
                                json.loads(av.get("ia_pontos_faltantes","[]") or "[]"),
                                av.get("ia_analise_detalhada"), av.get("ia_status")
                            )
                            # Registrar aprendizado
                            try:
                                db_module.registrar_aprendizado(
                                    pratica_num, sub_idx,
                                    pratica_nome, sub_nome,
                                    aud.get("unidade",""), aud.get("area",""), aud.get("ciclo",""),
                                    nota_sa or 0,
                                    int(float(av.get("ia_nota_sugerida") or 0)),
                                    av.get("ia_decisao",""),
                                    _obs_ap[:500],
                                    _dec_cons, _nota_ap,
                                    _resp_limpa[:300] if _resp_limpa else "Consenso via revisão colaborativa"
                                )
                            except Exception:
                                pass
                            st.success(f"✅ Consenso aplicado: **{_dec_cons}** | Nota **{_nota_ap}**")
                            _limpar_cache_auditoria()
                            # Limpar session_state dos widgets de decisão/NC/comentários
                            # para forçar Streamlit a reler os valores do banco
                            for _sk in [f"{kp}_dec", f"{kp}_desc", f"{kp}_coment"]:
                                st.session_state.pop(_sk, None)
                            # Preservar estado antes do rerun
                            st.session_state[f'prat_{pratica_num}_open'] = True
                            st.session_state[f'{kp}_open'] = True
                            st.session_state[chat_key] = True
                            st.rerun(scope="fragment")
                    with col_send:
                        if st.button("📨 Enviar para IA", key=f"{kp}_chat_send",
                                     type="primary", use_container_width=True,
                                     disabled=not nova_msg_input.strip()):
                            if not api_key:
                                st.error("Configure a chave da API OpenAI em ⚙️ Configurações.")
                            elif nova_msg_input.strip():
                                _chat_ph = st.empty()
                                _chat_ph.info("🤖 IA revisando análise com seu contexto adicional…")
                                try:
                                    from ai_analyzer import AuditAIAnalyzer
                                    # Usa ev_mapa_cache (mesmos arquivos exibidos na UI)
                                    files_chat = ev_mapa_cache.get((pratica_num, sub_idx), [])
                                    _eco = st.session_state.get('modo_ia','completo') == 'economico'
                                    analyzer_chat = AuditAIAnalyzer(api_key, economico=_eco)
                                    hist_openai = db_module.historico_para_openai(chat_msgs)
                                    exemplos_fw = db_module.buscar_exemplos_similares(pratica_num, sub_idx)

                                    resultado_revisao = analyzer_chat.revisar_com_contexto(
                                        pratica_num=pratica_num,
                                        subitem_idx=sub_idx,
                                        pratica_nome=pratica_nome,
                                        subitem_nome=sub_nome,
                                        evidencia_descricao=ev_desc,
                                        niveis_planilha={k: av.get(f"nivel_{k}", "") or "" for k in range(5)},
                                        nota_self_assessment=nota_sa or 0,
                                        evidence_files=files_chat,
                                        historico_chat=hist_openai,
                                        nova_mensagem=nova_msg_input.strip(),
                                        exemplos_aprendizado=exemplos_fw
                                    )

                                    resposta_ia = resultado_revisao["resposta"]
                                    mudou = resultado_revisao["mudou_decisao"]
                                    nova_dec_ia = resultado_revisao.get("nova_decisao")
                                    nova_nota_ia = resultado_revisao.get("nova_nota")
                                    nova_conf_ia = resultado_revisao.get("nova_confianca")

                                    # Salvar mensagem do usuário
                                    dec_atual_snap = av.get("ia_decisao")
                                    nota_atual_snap = av.get("ia_nota_sugerida")
                                    db_module.salvar_mensagem_chat(
                                        aud_id, pratica_num, sub_idx,
                                        role="user",
                                        conteudo=nova_msg_input.strip(),
                                        decisao_snapshot=dec_atual_snap,
                                        nota_snapshot=nota_atual_snap
                                    )

                                    # Salvar resposta da IA
                                    db_module.salvar_mensagem_chat(
                                        aud_id, pratica_num, sub_idx,
                                        role="assistant",
                                        conteudo=resposta_ia,
                                        decisao_snapshot=nova_dec_ia if nova_dec_ia else dec_atual_snap,
                                        nota_snapshot=nova_nota_ia if nova_nota_ia is not None else nota_atual_snap
                                    )

                                    # Se a IA revisou a decisão → atualizar banco
                                    # ── Resolver conflito nota ↔ decisão (a nota vence) ──
                                    if nova_nota_ia is not None:
                                        if nova_nota_ia == 0 and nova_dec_ia != "inexistente":
                                            nova_dec_ia = "inexistente"
                                            mudou = True
                                        elif nova_nota_ia > 0 and nova_nota_ia < (nota_sa or 5) and nova_dec_ia == "permanece":
                                            nova_dec_ia = "insuficiente"
                                            mudou = True
                                    if mudou and nova_dec_ia:
                                        nota_rev = nova_nota_ia if nova_nota_ia is not None else calcular_nota_final(nota_sa, nova_dec_ia)
                                        conf_rev = nova_conf_ia or av.get("ia_confianca", "media")
                                        # ── Limpar marcadores técnicos da resposta da IA ──
                                        import re as _re_auto
                                        _resp_ia_limpa = _re_auto.sub(
                                            r'REVIS[ÃA]O_[A-Z_ÇÃÁÉÍÓÚ]+:\s*\S+', "", resposta_ia
                                        ).strip()
                                        # ── Montar desc_nc: sempre salva justificativa IA ──
                                        _desc_nc_auto = _resp_ia_limpa[:1500] if _resp_ia_limpa else av.get("descricao_nc", "")
                                        # ── Comentário automático baseado na nova decisão ──
                                        if nova_dec_ia == "permanece":
                                            _coment_auto = "Nota permanece"
                                        elif nova_dec_ia == "inexistente":
                                            _coment_auto = "Nota passa para 0"
                                        elif nova_dec_ia == "insuficiente":
                                            _coment_auto = f"Nota passa para {nota_rev}" if nota_rev is not None else "Evidência insuficiente"
                                        else:
                                            _coment_auto = (
                                                f"Revisão colaborativa: {nova_msg_input.strip()[:300]}"
                                            )
                                        salvar_avaliacao(
                                            aud_id, pratica_num, pratica_nome,
                                            sub_idx, sub_nome, ev_desc,
                                            av.get("nivel_0",""), av.get("nivel_1",""),
                                            av.get("nivel_2",""), av.get("nivel_3",""),
                                            av.get("nivel_4",""), nota_sa,
                                            nova_dec_ia, nota_rev,
                                            _desc_nc_auto,
                                            _coment_auto,
                                            nova_dec_ia, nota_rev, conf_rev,
                                            json.loads(av.get("ia_pontos_atendidos","[]") or "[]"),
                                            json.loads(av.get("ia_pontos_faltantes","[]") or "[]"),
                                            resposta_ia, "ok"
                                        )
                                        st.success(f"✅ IA revisou a análise! Nova decisão: **{nova_dec_ia}** | Nota: **{nota_rev}**")
                                    else:
                                        st.info("ℹ️ IA manteve a análise anterior. Veja a resposta no chat.")

                                    # Limpar session_state dos widgets (inclui chat_input para limpar caixa)
                                    for _sk in [f"{kp}_dec", f"{kp}_desc", f"{kp}_coment", f"{kp}_chat_input"]:
                                        st.session_state.pop(_sk, None)
                                    # Preservar estado antes do rerun
                                    st.session_state[f'prat_{pratica_num}_open'] = True
                                    st.session_state[f'{kp}_open'] = True
                                    st.session_state[chat_key] = True
                                    _chat_ph.empty()
                                    _limpar_cache_auditoria()
                                    st.rerun(scope="fragment")

                                except Exception as e_chat:
                                    _chat_ph.empty()
                                    err_msg = str(e_chat)
                                    _mostrar_erro_openai(e_chat)


        # (Critérios e Evidências movidos para cima, antes do Checklist)

        # ── DECISÃO MANUAL ──
        st.markdown("**Decisão Final do Auditor:**")
        ia_sugestao=av.get('ia_decisao') if ia_status_val=='ok' else None
        decisao_default=decisao_atual if decisao_atual in ['permanece','insuficiente','inexistente','pendente'] else 'pendente'
        if ia_sugestao and decisao_atual=='pendente': decisao_default=ia_sugestao

        opcoes={'permanece':'✅ Nota permanece',
                'insuficiente':'⚠️ Evidência insuficiente (−1)',
                'inexistente':'❌ Evidência inexistente (→ 0)',
                'pendente':'⏳ Pendente'}
        idx_d=list(opcoes.keys()).index(decisao_default) if decisao_default in opcoes else 3
        decisao_sel=st.radio("dec",list(opcoes.keys()),
            format_func=lambda x:opcoes[x],index=idx_d,
            key=f"{kp}_dec",horizontal=True,label_visibility="collapsed")

        # Quando insuficiente: auditor escolhe nota de 0 até nota_sa-1
        nota_livre_sel = None
        if decisao_sel == "insuficiente" and nota_sa is not None and nota_sa > 0:
            _max_ins = max(0, nota_sa - 1)
            if _max_ins > 0:
                nota_livre_sel = st.select_slider(
                    "🔢 Nota final (escolha de 0 a {_max_ins}):".replace("{_max_ins}", str(_max_ins)),
                    options=list(range(0, _max_ins + 1)),
                    value=st.session_state.get(f"{kp}_nota_livre", _max_ins),
                    key=f"{kp}_nota_livre",
                    help=f"SA={nota_sa} → máximo possível={_max_ins}. Escolha a nota que a evidência justifica.",
                )
            else:
                nota_livre_sel = 0
                st.info("Nota SA=1 → evidência insuficiente leva a nota 0.")
        nota_calc = calcular_nota_final(nota_sa, decisao_sel, nota_livre_sel)
        if nota_calc is not None:
            em,ed=ESCALA.get(nota_calc,("",""))
            st.markdown(f"**Nota Final:** {badge_nota_html(nota_calc)}",unsafe_allow_html=True)

        # Pré-preencher: banco tem prioridade; IA só como sugestão inicial
        # (quando o usuário ainda não salvou manualmente, desc_nc fica vazia
        #  mas ia_analise_detalhada tem conteúdo → usar como sugestão)
        if desc_nc_atual:
            # Usuário já salvou algo → usar o que está no banco
            default_desc = desc_nc_atual
        elif ia_status_val == 'ok' and decisao_atual == (av.get('ia_decisao') or ''):
            # IA analisou mas usuário NÃO editou ainda → sugerir conteúdo da IA
            default_desc = av.get('ia_analise_detalhada', '') or ''
        else:
            # Usuário já salvou com desc vazia intencionalmente
            default_desc = ''

        # ── Comentário automático ──
        # Só auto-gerar se: (a) comentário atual é vazio ou auto-gerado
        # E (b) o usuário acabou de mudar a decisão no radio button
        _PADROES_AUTO = ("Nota permanece", "Nota passa para ", "Revisão colaborativa")
        _coment_eh_auto = (not coment_atual) or any(coment_atual.startswith(p) for p in _PADROES_AUTO)
        if _coment_eh_auto and decisao_sel != decisao_atual:
            _nota_prev_ck = calcular_nota_final(nota_sa, decisao_sel, nota_livre_sel)
            if decisao_sel == 'permanece':
                default_coment = "Nota permanece"
            elif decisao_sel == 'inexistente':
                default_coment = "Nota passa para 0"
            elif decisao_sel == 'insuficiente':
                default_coment = f"Nota passa para {_nota_prev_ck}" if _nota_prev_ck is not None else "Evidência insuficiente"
            else:
                default_coment = coment_atual or ""
        else:
            default_coment = coment_atual

        if decisao_sel in ('insuficiente','inexistente'):
            desc_nc=st.text_area("📝 Descrição da Não Conformidade",
                value=default_desc[:2000] if default_desc else '',
                key=f"{kp}_desc",height=220,
                placeholder="Descreva o que falta ou por que a evidência não é válida...")
        else:
            # Exibir (read-only) mesmo quando permanece, para consulta
            desc_nc=desc_nc_atual
            if desc_nc_atual:
                st.text_area("📝 Descrição da Não Conformidade (salva)",
                    value=desc_nc_atual[:2000],
                    height=180, disabled=True,
                    key=f"{kp}_desc_ro",
                    label_visibility="visible")

        coment=st.text_area("💬 Comentários",value=default_coment,key=f"{kp}_coment",
            height=90,
            placeholder="Observações, recomendações...")

        # Botão salvar
        if st.button("💾 Salvar Decisão",key=f"{kp}_save",type="primary"):
            salvar_avaliacao(
                aud_id,pratica_num,pratica_nome,sub_idx,sub_nome,ev_desc,
                av.get('nivel_0',''),av.get('nivel_1',''),av.get('nivel_2',''),
                av.get('nivel_3',''),av.get('nivel_4',''),nota_sa,
                decisao_sel,nota_calc,
                desc_nc if decisao_sel in ('insuficiente','inexistente') else '',coment,
                av.get('ia_decisao'),av.get('ia_nota_sugerida'),av.get('ia_confianca'),
                json.loads(av.get('ia_pontos_atendidos','[]') or '[]'),
                json.loads(av.get('ia_pontos_faltantes','[]') or '[]'),
                av.get('ia_analise_detalhada'),av.get('ia_status'))
            # Registrar aprendizado se houve chat e decisão diferiu da IA
            try:
                _msgs_aprendiz = db_module.carregar_chat(aud_id, pratica_num, sub_idx)
                if _msgs_aprendiz:
                    _obs_learn = " | ".join(m.get("conteudo", m.get("content", ""))[:200] for m in _msgs_aprendiz if m["role"]=="user")
                    db_module.registrar_aprendizado(
                        pratica_num, sub_idx,
                        pratica_nome, sub_nome,
                        aud.get("unidade",""), aud.get("area",""), aud.get("ciclo",""),
                        nota_sa or 0,
                        int(float(av.get("ia_nota_sugerida") or 0)),
                        av.get("ia_decisao",""),
                        _obs_learn[:500],
                        decisao_sel, nota_calc or 0,
                        (desc_nc or coment or "")[:300]
                    )
            except Exception:
                pass
            # Limpar apenas dec e nota_livre; preservar desc/coment para que
            # o re-render não recalcule defaults com conteúdo da IA
            st.session_state.pop(f"{kp}_dec", None)
            st.session_state.pop(f"{kp}_nota_livre", None)
            # Manter subitem aberto para o usuário ver o resultado
            st.session_state[f'{kp}_open'] = True
            st.session_state[f'prat_{pratica_num}_open'] = True
            _limpar_cache_auditoria()
            st.toast("✅ Decisão salva!", icon="✅"); st.rerun(scope="fragment")

if st.session_state.pagina == 'dashboard':
    st.title("🏠 Dashboard — Auditoria de Automação TA")
    df = listar_auditorias()
    total_areas = sum(len(v) for v in UNIDADES_AREAS.values())
    concluidas  = len(df[df['status']=='concluida']) if len(df)>0 else 0
    andamento   = len(df[df['status']=='em_andamento']) if len(df)>0 else 0
    # ia_analisados vem direto do SQL de listar_auditorias (sem N+1 queries)
    com_ia = int(df['ia_analisados'].gt(0).sum()) if len(df) > 0 and 'ia_analisados' in df.columns else 0

    c1,c2,c3,c4,c5 = st.columns(5)
    metrics = [("🏭 Total Áreas",total_areas),("✅ Concluídas",concluidas),
               ("🔄 Em Andamento",andamento),("🤖 Com Análise IA",com_ia),
               ("📈 Progresso",f"{round(concluidas/total_areas*100)}%")]
    for col,(label,val) in zip([c1,c2,c3,c4,c5],metrics):
        col.metric(label,val)

    st.markdown("---")
    if len(df)==0:
        st.info("📭 Nenhuma auditoria cadastrada. Use **➕ Nova Auditoria** para começar!")
    else:
        st.subheader("📋 Auditorias")
        for _,row in df.iterrows():
            c1,c2,c3,c4,c5 = st.columns([3,2,2,2,2])
            with c1: st.markdown(f"**{row['unidade']} / {row['area']}** — {row.get('ciclo','')}")
            _st_dash = row.get('status','em_andamento')
            _DASH_LABELS = {'em_andamento':'🔄 Em Andamento','concluida':'✅ Concluída',
                            'em_revisao':'🔍 Em Revisão','aprovada':'🏆 Aprovada','arquivada':'🗃️ Arquivada'}
            with c2: st.markdown(_DASH_LABELS.get(_st_dash,'🔄'))
            with c3: st.markdown(f"📅 {str(row.get('data_atualizacao',''))[:10]}")
            with c4:
                ia_ok = int(row.get('ia_analisados', 0) or 0)
                tot   = int(row.get('total_subitens', 0) or 0)
                st.markdown(f"🤖 {ia_ok}/{tot} IA")
            with c5:
                if st.button("Abrir",key=f"op_{row['id']}"):
                    st.session_state.auditoria_id=int(row['id'])
                    st.session_state.pagina='auditar'; st.rerun()

    st.markdown("---")
    st.subheader("🗺️ Cobertura por Unidade")
    # Auditoria ativa
    _active = get_auditoria(st.session_state.auditoria_id) if st.session_state.auditoria_id else None
    _act_unidade = _active['unidade'] if _active else None
    _act_area    = _active['area']    if _active else None
    for unidade, areas in UNIDADES_AREAS.items():
        aud_u = df[df['unidade']==unidade] if len(df)>0 else pd.DataFrame()
        conc  = list(aud_u[aud_u['status']=='concluida']['area']) if len(aud_u)>0 else []
        and_  = list(aud_u[aud_u['status']=='em_andamento']['area']) if len(aud_u)>0 else []
        st.markdown(f"**{unidade}**")
        badge_html = ""
        for area in areas:
            is_active = (unidade == _act_unidade and area == _act_area)
            if is_active:
                bg, border, icon = "#1565C0", "border:2px solid #90CAF9;", "📍"
            elif area in conc:
                bg, border, icon = "#388E3C", "", "✅"
            elif area in and_:
                bg, border, icon = "#E65100", "", "🔄"
            else:
                bg, border, icon = "#78909C", "", "⬜"
            badge_html += (
                f'<span title="{area}" style="display:inline-block;background:{bg};color:white;'
                f'padding:4px 10px;border-radius:8px;font-size:0.82em;margin:3px 4px;'
                f'{border}">{icon} {area}</span>'
            )
        st.markdown(badge_html, unsafe_allow_html=True)
        st.markdown("")

# ──────────────────────────────────────────────────────────────────────────────
# NOVA AUDITORIA
# ──────────────────────────────────────────────────────────────────────────────
elif st.session_state.pagina == 'nova':
    st.title("➕ Nova Auditoria")

    # Inicializar session_state para os campos (chaves dos widgets)
    if 'nova_assessment_input' not in st.session_state:
        st.session_state['nova_assessment_input'] = ''
    if 'nova_evidence_input' not in st.session_state:
        st.session_state['nova_evidence_input'] = ''

    # ── Unidade / Área / Ciclo ──
    c1, c2 = st.columns(2)
    with c1: unidade_sel = st.selectbox("🏭 Unidade", list(UNIDADES_AREAS.keys()), key='nova_unidade')
    with c2: area_sel = st.selectbox("🔧 Área", UNIDADES_AREAS.get(unidade_sel, []), key='nova_area')
    ciclo_sel = st.text_input("📅 Ciclo / Ano", placeholder="Ex: 2026", key='nova_ciclo')

    st.markdown("---")
    st.markdown("#### 📂 Arquivos")

    # Campo Assessment com botão de browser
    col_txt1, col_btn1 = st.columns([5, 1])
    with col_btn1:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("📂 Selecionar", key="btn_browse_assessment", use_container_width=True):
            path = browser_arquivo_xlsx("Selecionar Planilha de Assessment (.xlsx)", initial_dir=r"C:\AuditoriaTA\Self Assessment")
            if path:
                # Escrever diretamente na chave do widget — único jeito de atualizar o text_input
                st.session_state['nova_assessment_input'] = path
                st.rerun()
    with col_txt1:
        assessment_path = st.text_input(
            "Planilha de Assessment (.xlsx)",
            placeholder=r"C:\Auditorias\Assessment_Aciaria_JF.xlsx",
            key='nova_assessment_input'
        )

    # Preview assessment
    if assessment_path:
        existe = os.path.isfile(assessment_path)
        cor = "green" if existe else "red"
        icone = "✅" if existe else "❌"
        st.markdown(f"<small style='color:{cor}'>{icone} <code>{assessment_path}</code></small>", unsafe_allow_html=True)

    # Campo Pasta Evidências com botão de browser
    col_txt2, col_btn2 = st.columns([5, 1])
    with col_btn2:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("📂 Selecionar", key="btn_browse_evidence", use_container_width=True):
            path = browser_pasta("Selecionar Pasta Raiz das Evidências", initial_dir=r"C:\AuditoriaTA\Evidências")
            if path:
                # Escrever diretamente na chave do widget
                st.session_state['nova_evidence_input'] = path
                st.rerun()
    with col_txt2:
        evidence_folder = st.text_input(
            "Pasta Raiz das Evidências",
            placeholder=r"C:\Auditorias\Evidencias\JF\Aciaria",
            help="A pasta deve ter subpastas como '[1] ROTINA DE TA', '[2] SOBRESSALENTES', etc.",
            key='nova_evidence_input'
        )

    # Preview pasta
    if evidence_folder:
        existe = os.path.isdir(evidence_folder)
        cor = "green" if existe else "red"
        icone = "✅" if existe else "❌"
        st.markdown(f"<small style='color:{cor}'>{icone} <code>{evidence_folder}</code></small>", unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("#### 🤖 Configuração de IA")
    api_key = st.text_input(
        "Chave API OpenAI (opcional)", type="password",
        placeholder="sk-...",
        help="Necessário para análise automática com IA. Pode configurar depois em ⚙️ Configurações.",
        key='nova_api_key'
    )

    st.markdown("")
    submitted = st.button("🚀 Criar Auditoria", type="primary", key='btn_criar_auditoria')

    if submitted:
        unidade = st.session_state.get('nova_unidade', '')
        area = st.session_state.get('nova_area', '')
        ciclo = st.session_state.get('nova_ciclo', '')
        assessment_path = st.session_state.get('nova_assessment_input', '')
        evidence_folder = st.session_state.get('nova_evidence_input', '')
        api_key = st.session_state.get('nova_api_key', '')
        if not unidade or not area or not ciclo:
            st.error("Preencha unidade, área e ciclo.")
        elif assessment_path and not os.path.isfile(assessment_path):
            st.error(f"Arquivo não encontrado: {assessment_path}")
        else:
            aud_id=criar_auditoria(unidade,area,ciclo,assessment_path,evidence_folder,api_key)
            if aud_id:
                if assessment_path and os.path.isfile(assessment_path):
                    praticas=parse_assessment(assessment_path)
                    st.write(f"🔍 Diagnostic: {len(praticas)} práticas encontradas.")
                    for p in praticas:
                        total_sub = len(p['subitems'])
                        notas_encontradas = [s['nota_sa'] for s in p['subitems'] if s['nota_sa'] is not None]
                        st.write(f"  - Prática {p['num']}: {total_sub} subitens, notas: {notas_encontradas}")
                        for idx,sub in enumerate(p['subitems']):
                            salvar_avaliacao(aud_id,p['num'],p['nome'],idx,sub['nome'],sub['evidencia'],
                                sub['niveis'].get(0,''),sub['niveis'].get(1,''),sub['niveis'].get(2,''),
                                sub['niveis'].get(3,''),sub['niveis'].get(4,''),sub['nota_sa'],
                                'pendente',None,'','')
                st.session_state.auditoria_id=aud_id
                st.success("✅ Auditoria criada!")
                if st.button("Ir para Auditoria"):
                    st.session_state.pagina='auditar'; st.rerun()
                st.stop() # Force stop to let user see logs

# ──────────────────────────────────────────────────────────────────────────────
# AUDITAR
# ──────────────────────────────────────────────────────────────────────────────
elif st.session_state.pagina == 'auditar':
    if not st.session_state.auditoria_id:
        st.title("📋 Selecionar Auditoria")
        df=listar_auditorias()
        for _,row in df.iterrows():
            c1,c2,c3=st.columns([4,2,2])
            c1.markdown(f"**{row['unidade']} / {row['area']}** — {row.get('ciclo','')}")
            c2.markdown("✅" if row['status']=='concluida' else "🔄")
            if c3.button("Abrir",key=f"s_{row['id']}"):
                st.session_state.auditoria_id=int(row['id']); st.rerun()
        st.stop()

    aud_id=st.session_state.auditoria_id
    aud=get_auditoria(aud_id)
    if not aud: st.error("Auditoria não encontrada."); st.stop()

    st.title(f"📋 {aud['area']} — {aud['unidade']}")

    # ── Config rápida ──
    with st.expander("⚙️ Configurações da Auditoria", expanded=False):
        # Assessment .xlsx
        col_a1, col_a2 = st.columns([5, 1])
        with col_a2:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("📂", key="btn_cfg_ass", use_container_width=True, help="Selecionar arquivo"):
                path = browser_arquivo_xlsx("Selecionar Planilha Assessment (.xlsx)", initial_dir=r"C:\AuditoriaTA\Self Assessment")
                if path:
                    st.session_state['qa'] = path   # chave direta do widget
                    st.rerun()
        with col_a1:
            novo_ass = st.text_input(
                "Assessment .xlsx",
                value=aud.get('assessment_file_path','') or '',
                key="qa"
            )

        if novo_ass:
            ok = os.path.isfile(novo_ass)
            st.markdown(f"<small style='color:{'green' if ok else 'red'}'>{'✅' if ok else '❌'} <code>{novo_ass}</code></small>", unsafe_allow_html=True)

        # Pasta Evidências
        col_b1, col_b2 = st.columns([5, 1])
        with col_b2:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("📂", key="btn_cfg_ev", use_container_width=True, help="Selecionar pasta"):
                path = browser_pasta("Selecionar Pasta de Evidências", initial_dir=r"C:\AuditoriaTA\Evidências")
                if path:
                    st.session_state['qe'] = path   # chave direta do widget
                    st.rerun()
        with col_b1:
            novo_ev = st.text_input(
                "Pasta Evidências",
                value=aud.get('evidence_folder_path','') or '',
                key="qe"
            )

        if novo_ev:
            ok = os.path.isdir(novo_ev)
            st.markdown(f"<small style='color:{'green' if ok else 'red'}'>{'✅' if ok else '❌'} <code>{novo_ev}</code></small>", unsafe_allow_html=True)

        # API Key
        novo_key = st.text_input("API Key OpenAI", type="password",
                                  value=aud.get('openai_api_key','') or '', key="qk")
        cc1, cc2 = st.columns(2)
        with cc1:
            if st.button("💾 Salvar Config"):
                atualizar_config(aud_id, novo_ass, novo_ev, novo_key)
                _limpar_cache_auditoria()
                # Validar a chave recém-salva
                if novo_key and novo_key.strip().startswith('sk-'):
                    with st.spinner("🔑 Validando chave API..."):
                        try:
                            from ai_analyzer import validar_chave_openai
                            res = validar_chave_openai(novo_key.strip())
                            if res['valida']:
                                st.success("✅ Config salva e chave API validada com sucesso!")
                            else:
                                st.success("✅ Config salva!")
                                st.warning(res['erro'])
                        except Exception:
                            st.success("✅ Config salva!")
                else:
                    st.success("✅ Config salva!")
                st.rerun()
        with cc2:
            if st.button("🔄 Recarregar Assessment"):
                ap = aud.get('assessment_file_path', '')
                if ap and os.path.isfile(ap):
                    praticas = parse_assessment(ap)
                    for p in praticas:
                        for idx, sub in enumerate(p['subitems']):
                            salvar_avaliacao(aud_id, p['num'], p['nome'], idx, sub['nome'], sub['evidencia'],
                                sub['niveis'].get(0,''), sub['niveis'].get(1,''), sub['niveis'].get(2,''),
                                sub['niveis'].get(3,''), sub['niveis'].get(4,''), sub['nota_sa'],
                                'pendente', None, '', '')
                    _limpar_cache_auditoria()
                    st.success("Recarregado!"); st.rerun()
                else: st.error("Arquivo não encontrado.")

    df_aval=carregar_avaliacoes(aud_id)
    if len(df_aval)==0:
        st.warning("⚠️ Nenhum dado carregado. Configure o Assessment e recarregue.")
        st.stop()

    # ── Barra de progresso ──
    total_items=len(df_aval)
    avaliados=len(df_aval[df_aval['decisao']!='pendente'])
    ia_feitos=df_aval['ia_status'].eq('ok').sum()
    prog=avaliados/total_items if total_items>0 else 0
    st.progress(prog, text=f"Auditados: {avaliados}/{total_items} ({int(prog*100)}%)  |  🤖 IA: {ia_feitos}/{total_items}")

    # ── Botão ANÁLISE IA COMPLETA ──
    ev_folder=aud.get('evidence_folder_path','') or ''
    api_key=aud.get('openai_api_key','') or ''
    # Mapa de evidências construído UMA vez — usado por toda a tela de auditoria
    ev_mapa_cache = construir_mapa_evidencias(ev_folder)

    col_ia1, col_ia2 = st.columns([3,1])
    with col_ia1:
        if api_key:
            st.markdown('<span class="badge-ai">🤖 IA Disponível</span>', unsafe_allow_html=True)
        else:
            st.warning("⚠️ Configure sua chave API OpenAI para habilitar análise com IA.")
    with col_ia2:
        run_all_ia = st.button("🤖 Analisar TUDO com IA", type="primary", disabled=not api_key,
                               help="Analisa todos os subitens automaticamente com GPT-4o")

    if run_all_ia and api_key:
        try:
            from ai_analyzer import AuditAIAnalyzer
            _eco = st.session_state.get('modo_ia','completo') == 'economico'
            analyzer=AuditAIAnalyzer(api_key, economico=_eco)
            evidence_map=ev_mapa_cache  # usa o mesmo mapa exibido na UI

            subitems_data=[]
            for _,av in df_aval.iterrows():
                subitems_data.append({
                    'pratica_num':int(av['pratica_num']),'pratica_nome':av['pratica_nome'],
                    'subitem_idx':int(av['subitem_idx']),'subitem_nome':av['subitem_nome'],
                    'evidencia_descricao':av.get('evidencia_descricao','') or '',
                    'niveis':{k:av.get(f'nivel_{k}','') or '' for k in range(5)},
                    'nota_sa':av.get('nota_self_assessment')
                })

            progress_bar=st.progress(0)
            status_txt=st.empty()
            results=[]

            for i,sub in enumerate(subitems_data):
                key=(sub['pratica_num'],sub['subitem_idx'])
                files=evidence_map.get(key,[])
                _n_f = len(files)
                _n_imgs_b = sum(1 for f in files if f.suffix.lower() in {'.jpg','.jpeg','.png','.gif','.webp','.bmp'})
                _n_docs_b = _n_f - _n_imgs_b
                _arq_info = f" [{_n_docs_b} doc(s) + {_n_imgs_b} img(s)]" if _n_f > 0 else " [sem evidências]"
                status_txt.markdown(f"🔍 Analisando: **{sub['subitem_nome'][:60]}**...{_arq_info}")
                try:
                    res=analyzer.analyze_subitem(
                        pratica_num=sub['pratica_num'],
                        subitem_idx=sub['subitem_idx'],
                        pratica_nome=sub['pratica_nome'],
                        subitem_nome=sub['subitem_nome'],
                        evidencia_descricao=sub.get('evidencia_descricao',''),
                        niveis_planilha=sub.get('niveis',{}),
                        nota_self_assessment=sub.get('nota_sa',0) or 0,
                        evidence_files=files)
                    res['pratica_num']=sub['pratica_num']; res['subitem_idx']=sub['subitem_idx']
                    res['status']='ok'
                except Exception as e:
                    res={'pratica_num':sub['pratica_num'],'subitem_idx':sub['subitem_idx'],
                         'decisao':'pendente','nota_sugerida':sub.get('nota_sa'),
                         'confianca':'baixa','pontos_atendidos':[],'pontos_faltantes':[],
                         'descricao_nc':'','comentarios':f'Erro IA: {e}',
                         'analise_detalhada':'','status':'erro','erro':str(e)}
                results.append(res)
                progress_bar.progress((i+1)/len(subitems_data))

            # Salvar resultados
            for r in results:
                row_av=df_aval[(df_aval['pratica_num']==r['pratica_num'])&(df_aval['subitem_idx']==r['subitem_idx'])].iloc[0]
                nota_sa=row_av.get('nota_self_assessment')
                ia_dec=r.get('decisao','pendente')
                ia_nota=r.get('nota_sugerida')
                nota_final=calcular_nota_final(nota_sa,ia_dec)
                salvar_avaliacao(
                    aud_id, r['pratica_num'], row_av['pratica_nome'],
                    r['subitem_idx'], row_av['subitem_nome'],
                    row_av.get('evidencia_descricao','') or '',
                    row_av.get('nivel_0',''),row_av.get('nivel_1',''),row_av.get('nivel_2',''),
                    row_av.get('nivel_3',''),row_av.get('nivel_4',''),
                    nota_sa, ia_dec, nota_final,
                    r.get('descricao_nc',''), r.get('comentarios',''),
                    ia_dec, ia_nota, r.get('confianca'),
                    r.get('pontos_atendidos',[]), r.get('pontos_faltantes',[]),
                    r.get('analise_detalhada',''), r.get('status','erro'))

            status_txt.empty(); progress_bar.empty()
            n_ok  = sum(1 for r in results if r.get('status')=='ok')
            n_err = len(results) - n_ok
            _limpar_cache_auditoria()
            st.success(f"✅ {n_ok} subitens analisados pela IA! {f'({n_err} com erro)' if n_err else ''}")

            # ── PAINEL DE REVISÃO PÓS-ANÁLISE ──
            st.markdown("---")
            st.markdown("## 🔍 Revisão dos Resultados da IA")
            st.info("💡 Confira abaixo o que a IA sugeriu para cada sub-item. Você pode aceitar ou ajustar manualmente na área de cada prática abaixo.")

            DECISAO_ICONS = {'permanece':'✅','insuficiente':'⚠️','inexistente':'❌','pendente':'⏳'}
            CONF_CORES    = {'alta':'🟢','media':'🟡','baixa':'🔴'}

            for r in results:
                if r.get('status') != 'ok':
                    continue
                p_num = r['pratica_num']; s_idx = r['subitem_idx']
                # Buscar nome do sub-item
                row_av = df_aval[(df_aval['pratica_num']==p_num) & (df_aval['subitem_idx']==s_idx)]
                if row_av.empty: continue
                row_av = row_av.iloc[0]
                nome_sub  = row_av['subitem_nome']
                nota_sa   = row_av.get('nota_self_assessment')
                dec_ia    = r.get('decisao','pendente')
                nota_ia   = r.get('nota_sugerida')
                conf      = r.get('confianca','baixa')
                atend     = r.get('pontos_atendidos',[])
                falt      = r.get('pontos_faltantes',[])
                analise   = r.get('analise_detalhada','')

                icon_dec  = DECISAO_ICONS.get(dec_ia, '⏳')
                icon_conf = CONF_CORES.get(conf, '⚪')
                nota_final_ia = calcular_nota_final(nota_sa, dec_ia)

                with st.expander(
                    f"{icon_dec} **{p_num}.{s_idx+1} — {nome_sub[:55]}** │ SA: `{nota_sa}` → IA: `{nota_ia}` {icon_conf}",
                    expanded=False
                ):
                    c_rev1, c_rev2 = st.columns([1,1])
                    with c_rev1:
                        st.markdown(f"**Decisão sugerida:** {icon_dec} `{dec_ia.upper()}`")
                        st.markdown(f"**Nota SA:** `{nota_sa}` → **Nota final sugerida:** `{nota_final_ia}`")
                        st.markdown(f"**Confiança:** {icon_conf} `{conf.upper()}`")
                        if atend:
                            st.markdown("**✅ Pontos atendidos:**")
                            for pt in atend: st.caption(f"  ✔ {pt}")
                        if falt:
                            st.markdown("**❌ Pontos faltantes:**")
                            for pt in falt: st.caption(f"  ✖ {pt}")
                    with c_rev2:
                        if analise:
                            st.markdown("**🔍 Análise detalhada:**")
                            st.write(analise)
            st.markdown("---")
            st.rerun()
        except ImportError as e:
            st.error(f"Dependência não instalada: {e}. Execute: pip install openai pdfplumber python-docx")
        except Exception as e_all_ia:
            _mostrar_erro_openai(e_all_ia)

    st.markdown("---")
    # df_aval já está atualizado via cache — não precisa recarregar aqui



    # ── Carregar TODOS os chats da auditoria de uma vez (1 query ao invés de N) ──
    _todos_chats = {}
    try:
        _todos_chats = db_module.carregar_todos_chats_auditoria(aud_id)
    except Exception:
        pass  # fallback: carregar_chat individual será usado

    # ── Iteração por prática ──
    for pratica_num in sorted(int(x) for x in df_aval['pratica_num'].dropna().unique()):
        df_p=df_aval[df_aval['pratica_num']==pratica_num]
        pratica_nome=df_p['pratica_nome'].iloc[0]
        msa=df_p['nota_self_assessment'].mean()
        mf=df_p['nota_final'].mean() if df_p['nota_final'].notna().any() else None
        pend=len(df_p[df_p['decisao']=='pendente'])
        ia_ok_p=df_p['ia_status'].eq('ok').sum()

        label=f"**{pratica_num}. {pratica_nome}**  |  SA: `{round(msa,1)}`  →  Final: `{round(mf,1) if mf is not None else '⏳'}`"
        label += f"  {'✅' if pend==0 else f'⏳ {pend} pend.'}  {'🤖 IA' if ia_ok_p>0 else ''}"

        _prat_key = f"prat_{pratica_num}_open"
        _prat_exp_open = st.session_state.get(_prat_key, False)
        with st.expander(label, expanded=_prat_exp_open):
            for _,av in df_p.iterrows():
                sub_idx=int(av['subitem_idx'])
                sub_nome=av['subitem_nome']
                nota_sa=av['nota_self_assessment']
                decisao_atual=av['decisao'] or 'pendente'
                desc_nc_atual=av.get('descricao_nc','') or ''
                coment_atual=av.get('comentarios','') or ''
                ev_desc=av.get('evidencia_descricao','') or ''
                ia_status_val=av.get('ia_status')
                kp=f"p{pratica_num}_s{sub_idx}"

                # ── Subitem expansível ──
                _dec_icon_sub = {"permanece": "✅", "insuficiente": "⚠️", "inexistente": "❌", "pendente": "⏳"}.get(decisao_atual, "⏳")
                _nota_sa_txt = f" | SA: {nota_sa}" if nota_sa is not None else ""
                _nota_final_val = av.get('nota_final')
                _nota_final_txt = f" → Final: {int(_nota_final_val)}" if pd.notna(_nota_final_val) else ""
                _sub_exp_label = f"{_dec_icon_sub} **{pratica_num}.{sub_idx+1}** — {sub_nome}{_nota_sa_txt}{_nota_final_txt}"
                # Manter subitem aberto se foi interagido recentemente
                _sub_was_open = st.session_state.get(f"{kp}_open", (decisao_atual == 'pendente'))
                with st.expander(_sub_exp_label, expanded=_sub_was_open):
                    _render_subitem_frag(aud_id, pratica_num, pratica_nome, sub_idx, api_key, aud)


    # ─────────────────────────────────────────────
    # ── BANNER DE STATUS ──────────────────────────
    _status_atual = aud.get('status','em_andamento')
    _STATUS_LABELS = {
        'em_andamento': ('🔄','Em Andamento','#1E3A5F','#90CAF9'),
        'concluida':    ('✅','Concluída – Aguardando Aprovação','#1B4332','#52B788'),
        'em_revisao':   ('🔍','Em Revisão','#4A2C00','#FFC107'),
        'aprovada':     ('🏆','Aprovada e Exportável','#1A1A2E','#A8EDEA'),
        'arquivada':    ('🗃️','Arquivada','#3A3A3A','#AAAAAA'),
    }
    _ico, _lbl, _bg, _fg = _STATUS_LABELS.get(_status_atual, ('','','#222','#fff'))
    st.markdown(f"""
    <div style='background:{_bg};border-radius:8px;padding:10px 18px;margin-bottom:10px;
                display:flex;align-items:center;gap:12px;'>
        <span style='font-size:1.5em;'>{_ico}</span>
        <span style='color:{_fg};font-weight:bold;font-size:1em;'>Status: {_lbl}</span>
    </div>""", unsafe_allow_html=True)

    st.markdown("---")
    c1,c2,c3,c4 = st.columns([3,1,1,1])
    with c2:
        if _status_atual == 'em_andamento':
            if st.button("🏁 Concluir", type="primary", help="Marca a auditoria como concluída e envia para aprovação"):
                db_module.atualizar_status(aud_id, 'concluida')
                _limpar_cache_auditoria()
                st.success("✅ Auditoria concluída! Aguardando aprovação."); st.rerun()
        elif _status_atual in ('concluida','em_revisao'):
            if st.button("✅ Aprovar", type="primary", help="Aprova a auditoria — planilha estará disponível para exportação"):
                db_module.atualizar_status(aud_id, 'aprovada')
                _limpar_cache_auditoria()
                st.success("🏆 Auditoria aprovada! Disponível para exportação."); st.rerun()
    with c3:
        if _status_atual in ('concluida','aprovada'):
            if st.button("🔍 Revisar", help="Reabre a auditoria para ajustes — status volta para 'Em Revisão'"):
                db_module.atualizar_status(aud_id, 'em_revisao')
                _limpar_cache_auditoria()
                st.info("🔍 Auditoria reaberta para revisão."); st.rerun()
        elif _status_atual == 'em_revisao':
            if st.button("🏁 Concluir Revisão", type="primary", help="Conclui a revisão"):
                db_module.atualizar_status(aud_id, 'concluida')
                st.success("✅ Revisão concluída! Aguardando aprovação."); st.rerun()
    with c4:
        if _status_atual != 'em_andamento':
            if st.button("🔓 Reabrir", help="Retorna para Em Andamento para edições completas"):
                db_module.atualizar_status(aud_id, 'em_andamento')
                _limpar_cache_auditoria()
                st.info("🔄 Auditoria reaberta para edição."); st.rerun()

# ──────────────────────────────────────────────────────────────────────────────
# RELATÓRIOS
# ──────────────────────────────────────────────────────────────────────────────
elif st.session_state.pagina == 'relatorios':
    st.title("📊 Relatórios")
    df=listar_auditorias()
    if len(df)==0: st.info("Nenhuma auditoria."); st.stop()

    aud_sel=st.selectbox("Selecionar Auditoria",df['id'].tolist(),
        format_func=lambda i: f"{df[df['id']==i]['unidade'].values[0]} / {df[df['id']==i]['area'].values[0]} — {df[df['id']==i]['ciclo'].values[0]}")

    if aud_sel:
        aud=get_auditoria(aud_sel); df_aval=carregar_avaliacoes(aud_sel)
        if len(df_aval)==0: st.warning("Sem dados."); st.stop()

        st.subheader(f"{aud['unidade']} — {aud['area']} | {aud.get('ciclo','')}")

        # Métricas IA
        ia_ok=df_aval['ia_status'].eq('ok').sum()
        ia_tot=len(df_aval)
        c1,c2,c3,c4=st.columns(4)
        c1.metric("Subitens totais",ia_tot)
        c2.metric("🤖 Analisados por IA",ia_ok)
        c3.metric("✅ Aprovados (permanece)",len(df_aval[df_aval['decisao']=='permanece']))
        c4.metric("⚠️+❌ Não conformidades",len(df_aval[df_aval['decisao'].isin(['insuficiente','inexistente'])]))

        # Tabela resumo
        resumo=[]
        for pn,grp in df_aval.groupby('pratica_num'):
            msa=grp['nota_self_assessment'].mean()
            mf=grp['nota_final'].mean() if grp['nota_final'].notna().any() else None
            ia_n=grp['ia_nota_sugerida'].mean() if grp['ia_nota_sugerida'].notna().any() else None
            resumo.append({'Prática':f"{pn}. {grp['pratica_nome'].iloc[0]}",'Subs':len(grp),
                'Média SA':round(msa,2) if pd.notna(msa) else '-',
                'Média Final':round(mf,2) if mf is not None and not pd.isna(mf) else '⏳',
                'Δ':round(mf-msa,2) if (mf is not None and not pd.isna(mf) and pd.notna(msa)) else '-',
                '🤖 Média IA':round(ia_n,2) if ia_n is not None and not pd.isna(ia_n) else '-',
                'Pend.':len(grp[grp['decisao']=='pendente'])})
        st.dataframe(pd.DataFrame(resumo),use_container_width=True,hide_index=True)

        with st.expander("Ver detalhes completos"):
            cols=['pratica_nome','subitem_nome','nota_self_assessment','decisao','nota_final',
                  'ia_decisao','ia_nota_sugerida','ia_confianca','descricao_nc','comentarios']
            df_show=df_aval[cols].copy()
            df_show.columns=['Prática','Subitem','SA','Decisão','Final','IA Decisão','IA Nota','IA Conf.','Desc. NC','Comentários']
            st.dataframe(df_show,use_container_width=True,hide_index=True)

        st.markdown("---")
        c1,c2=st.columns(2)
        with c1:
            st.markdown("#### 📥 Exportar Excel")
            _st_rel = aud.get('status','em_andamento')
            if _st_rel == 'aprovada':
                excel=gerar_excel(aud_sel)
                nome=f"Analise_{aud['area'].replace(' ','_')}_{aud['unidade'].replace(' ','_')}_{aud.get('ciclo','')}.xlsx"
                st.download_button("⬇️ Baixar Planilha (.xlsx)",excel,nome,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",type="primary")
            else:
                st.warning("⚠️ Exportação bloqueada — aprove a auditoria primeiro.")
                st.caption("Acesse **📁 Histórico de Auditorias** → clique em **🏆 Aprovar** para liberar.")
                if st.button("🏆 Aprovar esta auditoria agora", key="rel_aprovar_btn"):
                    if _st_rel in ('em_andamento','em_revisao'):
                        db_module.atualizar_status(aud_sel, 'concluida')
                    db_module.atualizar_status(aud_sel, 'aprovada')
                    _limpar_cache_auditoria()
                    st.success("🏆 Auditoria aprovada! Recarregue para exportar."); st.rerun()
        with c2:
            st.markdown("#### 📋 Exportar CSV")
            csv=pd.DataFrame(resumo).to_csv(index=False).encode('utf-8-sig')
            st.download_button("⬇️ Baixar CSV",csv,f"Resumo_{aud['area']}.csv","text/csv")

# ──────────────────────────────────────────────────────────────────────────────
# HISTÓRICO DE AUDITORIAS
# ──────────────────────────────────────────────────────────────────────────────
elif st.session_state.pagina == 'historico':
    st.title("📁 Histórico de Auditorias")
    st.markdown("Visualize, edite e aprove todas as auditorias — em andamento ou concluídas.")
    st.markdown("---")

    df_all = listar_auditorias()
    if len(df_all) == 0:
        st.info("Nenhuma auditoria cadastrada ainda.")
        st.stop()

    # ── Filtros ──────────────────────────────────────────────────────────────
    _STATUS_OPC = {
        "Todos":        None,
        "🔄 Em Andamento": "em_andamento",
        "✅ Concluída":    "concluida",
        "🔍 Em Revisão":   "em_revisao",
        "🏆 Aprovada":     "aprovada",
        "🗃️ Arquivada":    "arquivada",
    }
    col_f1, col_f2, col_f3 = st.columns([2,2,2])
    with col_f1:
        _unidades_disp = ["Todas"] + sorted(df_all['unidade'].dropna().unique().tolist())
        _fil_unidade = st.selectbox("Unidade", _unidades_disp, key="hist_fil_unidade")
    with col_f2:
        _fil_status_label = st.selectbox("Status", list(_STATUS_OPC.keys()), key="hist_fil_status")
        _fil_status = _STATUS_OPC[_fil_status_label]
    with col_f3:
        _fil_ciclo = st.text_input("Ciclo (filtro)", placeholder="ex: 2025", key="hist_fil_ciclo")

    df_fil = df_all.copy()
    if _fil_unidade != "Todas":
        df_fil = df_fil[df_fil['unidade'] == _fil_unidade]
    if _fil_status:
        df_fil = df_fil[df_fil['status'] == _fil_status]
    if _fil_ciclo.strip():
        df_fil = df_fil[df_fil['ciclo'].astype(str).str.contains(_fil_ciclo.strip(), case=False, na=False)]

    st.markdown(f"**{len(df_fil)} auditoria(s) encontrada(s)**")
    st.markdown("---")

    # ── Legenda de status ──────────────────────────────────────────────────
    _STATUS_INFO = {
        'em_andamento': ('🔄', '#90CAF9', 'Em Andamento'),
        'concluida':    ('✅', '#52B788', 'Concluída'),
        'em_revisao':   ('🔍', '#FFC107', 'Em Revisão'),
        'aprovada':     ('🏆', '#A8EDEA', 'Aprovada'),
        'arquivada':    ('🗃️', '#AAAAAA', 'Arquivada'),
    }

    if len(df_fil) == 0:
        st.warning("Nenhuma auditoria encontrada com os filtros selecionados.")
    else:
        for _, row in df_fil.iterrows():
            _st = row.get('status','em_andamento')
            _ico, _cor, _label = _STATUS_INFO.get(_st, ('','#888','Desconhecido'))
            _pend   = int(row.get('total_subitens',0) or 0) - int(row.get('subitens_avaliados',0) or 0)
            _media  = row.get('media_nota_final')
            _media_txt = f"📊 Média: {_media:.2f}" if _media is not None and not pd.isna(_media) else "📊 Sem notas"
            _data   = str(row.get('data_atualizacao',''))[:10]

            with st.container():
                st.markdown(f"""
                <div style='border:1px solid #3D5068;border-radius:10px;padding:14px 18px;
                            margin-bottom:8px;background:#1C2B3A;'>
                    <div style='display:flex;justify-content:space-between;align-items:center;'>
                        <div>
                            <span style='font-size:1.1em;font-weight:bold;color:#E0E0E0;'>
                                🏭 {row['unidade']} &nbsp;/&nbsp; {row['area']}
                            </span>
                            &nbsp;&nbsp;
                            <span style='font-size:0.85em;color:#90CAF9;'>Ciclo {row.get('ciclo','')}</span>
                        </div>
                        <div>
                            <span style='background:{_cor}22;color:{_cor};border:1px solid {_cor}55;
                                         padding:3px 10px;border-radius:20px;font-size:0.82em;font-weight:bold;'>
                                {_ico} {_label}
                            </span>
                        </div>
                    </div>
                    <div style='margin-top:6px;color:#90A4AE;font-size:0.82em;'>
                        📅 Atualizado: {_data} &nbsp;|&nbsp;
                        📋 {row.get('subitens_avaliados',0)}/{row.get('total_subitens',0)} subitens avaliados &nbsp;|&nbsp;
                        ⏳ {_pend} pendentes &nbsp;|&nbsp;
                        {_media_txt}
                    </div>
                </div>""", unsafe_allow_html=True)

                # Botões de ação
                _ba1, _ba2, _ba3, _ba4, _ba5 = st.columns([1.5, 1.2, 1.2, 1.2, 1.5])
                _rid = int(row['id'])

                with _ba1:
                    if st.button("✏️ Abrir / Editar", key=f"hist_abrir_{_rid}",
                                 help="Abre a auditoria na página de Auditar"):
                        # Se concluída/aprovada, reabre para revisão automaticamente
                        if _st in ('concluida','aprovada'):
                            db_module.atualizar_status(_rid, 'em_revisao')
                        _limpar_cache_auditoria()
                        st.session_state.auditoria_id = _rid
                        st.session_state.pagina = 'auditar'
                        st.rerun()

                with _ba2:
                    if _st in ('em_andamento','em_revisao'):
                        if st.button("🏁 Concluir", key=f"hist_concluir_{_rid}"):
                            db_module.atualizar_status(_rid, 'concluida')
                            _limpar_cache_auditoria()
                            st.success(f"✅ {row['area']} concluída!"); st.rerun()

                with _ba3:
                    if _st in ('concluida','em_revisao'):
                        if st.button("🏆 Aprovar", key=f"hist_aprovar_{_rid}",
                                     help="Aprova para exportação"):
                            db_module.atualizar_status(_rid, 'aprovada')
                            _limpar_cache_auditoria()
                            st.success(f"🏆 {row['area']} aprovada!"); st.rerun()

                with _ba4:
                    if _st == 'aprovada':
                        _nome_exp = f"Analise_{row['area'].replace(' ','_')}_{row['unidade'].replace(' ','_')}_{row.get('ciclo','')}.xlsx"
                        _excel_key = f"_excel_ready_{_rid}"
                        # Geração lazy: só gera quando clicado, armazena no session_state
                        if _excel_key in st.session_state:
                            st.download_button(
                                "⬇️ Baixar Excel",
                                st.session_state[_excel_key],
                                _nome_exp,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"hist_export_{_rid}",
                                type="primary",
                            )
                        else:
                            if st.button("⬇️ Preparar Excel", key=f"hist_prep_{_rid}",
                                         type="primary", help="Clique para gerar o arquivo Excel"):
                                with st.spinner("Gerando planilha..."):
                                    st.session_state[_excel_key] = gerar_excel(_rid)
                                st.rerun()
                    elif _st in ('concluida','em_revisao'):
                        st.button("⬇️ Exportar", key=f"hist_exp_dis_{_rid}",
                                  disabled=True, help="Aprove a auditoria para liberar exportação")

                with _ba5:
                    if _st != 'em_andamento':
                        if st.button("🔓 Reabrir", key=f"hist_reabrir_{_rid}",
                                     help="Retorna para Em Andamento"):
                            db_module.atualizar_status(_rid, 'em_andamento')
                            _limpar_cache_auditoria()
                            st.info(f"🔄 {row['area']} reaberta."); st.rerun()

                st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

    # ── Legenda visual ─────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("**Legenda de status:**")
    leg_cols = st.columns(5)
    for col, (k,v) in zip(leg_cols, _STATUS_INFO.items()):
        col.markdown(f"<span style='color:{v[1]};font-weight:bold;'>{v[0]} {v[2]}</span>", unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("""
    **Fluxo de aprovação:**
    ```
    ➕ Nova  →  🔄 Em Andamento  →  ✅ Concluída  →  🏆 Aprovada  →  ⬇️ Exportar
                                         ↑↓ 🔍 Em Revisão (para ajustes)
    ```
    > A exportação da planilha Excel só é liberada após **Aprovação**.
    """)


# ──────────────────────────────────────────────────────────────────────────────
# CONFIGURAÇÕES
# ──────────────────────────────────────────────────────────────────────────────
elif st.session_state.pagina == 'config':
    st.title("⚙️ Configurações")
    st.markdown("---")

    st.subheader("🤖 Sobre a Análise com IA")
    st.markdown("""
    O sistema utiliza **GPT-4o** ou **GPT-4o-mini** da OpenAI para analisar automaticamente as evidências de auditoria.

    **Como funciona:**
    1. O sistema lê todos os arquivos da pasta de evidências
    2. Mapeia automaticamente cada arquivo ao subitem correspondente (pela numeração das pastas: `1.1`, `1.2`, etc.)
    3. Para cada subitem, o modelo recebe:
       - 📸 Imagens (análise visual direta)
       - 📄 PDFs (extração de texto)
       - 📝 Documentos Word (extração de texto)
       - 📋 Os critérios exatos para cada nível de nota (0-4)
    4. A IA retorna: decisão, nota sugerida, pontos atendidos, pontos faltantes e análise detalhada
    5. O auditor **revisa e confirma** (ou altera) o parecer da IA
    """)

    # ── SELEÇÃO DE MODO DE ANÁLISE ──
    st.subheader("💰 Modo de Análise")
    _modo_atual = st.session_state.get('modo_ia', 'completo')
    _modo_sel = st.radio(
        "Escolha o modo:",
        options=['completo', 'economico'],
        format_func=lambda x: "🔍 Completo (GPT-4o — alta qualidade, maior custo)" if x == 'completo'
                              else "💲 Econômico (GPT-4o-mini — análise sólida, ~10× mais barato)",
        index=0 if _modo_atual == 'completo' else 1,
        key='radio_modo_ia',
        help="Completo: processa todos os documentos e imagens em alta resolução. Econômico: limita páginas e imagens por subitem."
    )
    if _modo_sel != _modo_atual:
        st.session_state['modo_ia'] = _modo_sel
        st.rerun()

    if _modo_sel == 'completo':
        st.info("🔍 **Modo Completo — GPT-4o**\n- Todos os documentos lidos integralmente\n- Até 20 imagens em alta resolução\n- 3500 tokens de resposta\n- Custo estimado por subitem: ~$0.05–$0.20 USD")
    else:
        st.success("💲 **Modo Econômico — GPT-4o-mini**\n- Máx 80.000 chars de texto total por subitem\n- Máx 6 imagens em baixa resolução\n- Máx 8 páginas por PDF\n- 1500 tokens de resposta\n- Custo estimado por subitem: ~$0.003–$0.015 USD (≈10× mais barato)")

    st.subheader("🔑 Configurar API Key Global")
    if st.session_state.auditoria_id:
        aud=get_auditoria(st.session_state.auditoria_id)
        current_key=aud.get('openai_api_key','') or ''
        new_key=st.text_input("Chave API OpenAI",type="password",value=current_key,
            placeholder="sk-proj-...", key="cfg_api_key_input")
        col_sv, col_test = st.columns([1,1])
        with col_sv:
            if st.button("💾 Salvar Chave API", key="btn_salvar_api_cfg"):
                atualizar_config(st.session_state.auditoria_id,
                                 aud.get('assessment_file_path',''),
                                 aud.get('evidence_folder_path',''), new_key)
                _limpar_cache_auditoria()
                st.success("✅ Chave salva com sucesso!")
                st.rerun()
        with col_test:
            if st.button("🔍 Testar Chave", key="btn_testar_api_cfg"):
                chave_test = new_key or current_key
                if not chave_test:
                    st.warning("Cole a chave antes de testar.")
                else:
                    with st.spinner("Testando conexão com OpenAI..."):
                        try:
                            from ai_analyzer import validar_chave_openai
                            res = validar_chave_openai(chave_test.strip())
                            if res['valida']:
                                st.success("✅ Chave válida! Conexão com OpenAI confirmada.")
                            else:
                                st.error(res['erro'])
                                if res['tipo_erro'] == 'sem_credito':
                                    st.info(
                                        "💡 **Como resolver:**\n"
                                        "1. Acesse [platform.openai.com](https://platform.openai.com)\n"
                                        "2. Vá em **Settings → seu Projeto → Limits**\n"
                                        "3. Defina um **Monthly budget** (ex: $20)\n"
                                        "4. Se o saldo for $0, vá em **Billing → Add to credit balance** e adicione crédito"
                                    )
                        except Exception as ex:
                            st.error(f"Erro ao testar: {ex}")
    else:
        st.info("Selecione uma auditoria primeiro.")

    st.markdown("---")
    st.subheader("📂 Estrutura esperada de pastas de evidências")
    st.code("""
📁 Aciaria (pasta raiz selecionada)
└── 📁 AUDITORIA AUTOMAÇÃO 2025    ← subpasta (nome livre)
    ├── 📁 1- ROTINA DE TA
    │   ├── 📁 1.1 ROTINA PERIÓDICO E POR EVENTO   ← detectado como P1, S1
    │   │   ├── screenshot1.jpg
    │   │   └── screenshot2.jpg
    │   ├── 📁 1.2 REDUNDÂNCIA E ORGANIZAÇÃO       ← detectado como P1, S2
    │   │   ├── backup_rede.jpg
    │   │   └── procedimento.docx
    │   └── 📁 1.3. TESTE BACKUP EM PLCS           ← detectado como P1, S3
    │       └── TESTE BACKUP DEZ25.pdf
    ├── 📁 2- SOBRESSALENTES
    │   ├── 📁 2.1 VERIFICAÇÃO DE SOBRESSALENTES   ← detectado como P2, S1
    │   └── 📁 2.2 EQUIPAMENTOS POSSUEM SOBRESS.   ← detectado como P2, S2
    └── ...
    """, language="")
    st.info("💡 O sistema detecta automaticamente as pastas que começam com 'X.Y' e mapeia ao subitem correto.")

# ──────────────────────────────────────────────────────────────────────────────
# PÁGINA: DIÁRIO DE AUDITORIA (Visão Panorâmica)
# ──────────────────────────────────────────────────────────────────────────────
elif st.session_state.pagina == 'diario':
    # ── PÁGINA: DIÁRIO DE AUDITORIA ──────────────────────────────────────────
    from diario_page import render_diario_page
    render_diario_page(
        listar_auditorias=listar_auditorias,
        get_auditoria=get_auditoria,
        carregar_avaliacoes=carregar_avaliacoes
    )

elif st.session_state.pagina == 'dados':
    st.title("🗄️ Gestão de Dados & Histórico")
    st.markdown("Gerencie o histórico de ciclos, backups automáticos e compare auditorias entre anos.")
    st.markdown("---")

    # ── TABS principais ──
    tab_hist, tab_comp, tab_backup, tab_io, tab_log, tab_learn = st.tabs([
        "📅 Histórico de Ciclos",
        "📈 Comparativo entre Ciclos",
        "💾 Backups",
        "📦 Exportar / Importar",
        "🔍 Audit Log",
        "🧠 Aprendizados IA",
    ])

    # ════════════════════════════════════════════════════════════
    # TAB 1 — Histórico de Ciclos
    # ════════════════════════════════════════════════════════════
    with tab_hist:
        st.subheader("📅 Todas as Auditorias por Ciclo")

        df_all = listar_auditorias()
        if df_all.empty:
            st.info("Nenhuma auditoria cadastrada ainda.")
        else:
            # Estilizar tabela
            STATUS_ICON = {"em_andamento": "🔄", "concluida": "✅", "arquivada": "📦"}
            df_show = df_all[["id","unidade","area","ciclo","status",
                              "data_criacao","data_atualizacao",
                              "total_subitens","subitens_avaliados","media_nota_final"]].copy()
            df_show["status"] = df_show["status"].map(lambda s: f"{STATUS_ICON.get(s,'❓')} {s}")
            df_show.columns = ["ID","Unidade","Área","Ciclo","Status",
                               "Criação","Atualização",
                               "Total Sub","Avaliados","Média Final"]
            df_show["Média Final"] = df_show["Média Final"].apply(
                lambda v: f"{v:.2f}" if pd.notna(v) else "—")
            st.dataframe(df_show, use_container_width=True, hide_index=True)

            st.markdown("---")
            # ── Ações por auditoria ──
            st.subheader("⚡ Ações rápidas")
            col_a, col_b, col_c = st.columns(3)

            with col_a:
                st.markdown("**Alterar status**")
                aud_ids   = df_all["id"].tolist()
                aud_nomes = [f"[{r.id}] {r.unidade} / {r.area} — {r.ciclo}"
                             for r in df_all.itertuples()]
                sel_status = st.selectbox("Auditoria", aud_nomes, key="sel_status_aud")
                sel_idx    = aud_nomes.index(sel_status)
                sel_id     = aud_ids[sel_idx]
                novo_status = st.selectbox("Novo status",
                    ["em_andamento","concluida","arquivada"], key="novo_status_val")
                if st.button("✅ Aplicar status", key="btn_status"):
                    db_module.atualizar_status(sel_id, novo_status)
                    _limpar_cache_auditoria()
                    st.success(f"Status atualizado para **{novo_status}**!")
                    st.rerun()

            with col_b:
                st.markdown("**Iniciar novo ciclo (duplicar)**")
                sel_dup = st.selectbox("Base", aud_nomes, key="sel_dup")
                sel_dup_idx = aud_nomes.index(sel_dup)
                sel_dup_id  = aud_ids[sel_dup_idx]
                ano_atual   = datetime.now().year
                novo_ciclo  = st.text_input("Nome do novo ciclo",
                    value=str(ano_atual), key="novo_ciclo_val",
                    help="Ex: 2026 ou '2026 — Auditoria Anual'")
                if st.button("🔁 Duplicar para novo ciclo", key="btn_dup"):
                    if not novo_ciclo.strip():
                        st.error("Informe o nome do ciclo.")
                    else:
                        with st.spinner("Duplicando auditoria…"):
                            novo_id = db_module.duplicar_auditoria(sel_dup_id, novo_ciclo.strip())
                        if novo_id:
                            st.success(f"✅ Nova auditoria criada! ID={novo_id} — Ciclo: {novo_ciclo}")
                            st.session_state.auditoria_id = novo_id
                            st.info("💡 Use **📋 Auditar** para iniciar o novo ciclo.")
                            st.rerun()
                        else:
                            st.error("Falha ao duplicar. Verifique se o ciclo já existe.")

            with col_c:
                st.markdown("**Excluir auditoria**")
                st.warning("⚠️ Operação irreversível. Um backup será criado automaticamente.")
                sel_del = st.selectbox("Auditoria a excluir", aud_nomes, key="sel_del")
                sel_del_idx = aud_nomes.index(sel_del)
                sel_del_id  = aud_ids[sel_del_idx]
                confirm_del = st.checkbox("Confirmo que desejo excluir", key="confirm_del")
                if st.button("🗑️ Excluir", key="btn_del", disabled=not confirm_del):
                    db_module.excluir_auditoria(sel_del_id)
                    if st.session_state.auditoria_id == sel_del_id:
                        st.session_state.auditoria_id = None
                    _limpar_cache_auditoria()
                    st.success("Auditoria excluída. Backup criado em dados/backups/.")
                    st.rerun()

    # ════════════════════════════════════════════════════════════
    # TAB 2 — Comparativo entre Ciclos
    # ════════════════════════════════════════════════════════════
    with tab_comp:
        st.subheader("📈 Comparativo de Notas entre Dois Ciclos")
        st.markdown("Selecione dois ciclos da mesma unidade/área para ver a evolução de cada subitem.")

        df_all2 = listar_auditorias()
        if len(df_all2) < 2:
            st.info("São necessárias pelo menos 2 auditorias para comparar.")
        else:
            aud_nomes2 = [f"[{r.id}] {r.unidade} / {r.area} — {r.ciclo}"
                         for r in df_all2.itertuples()]
            col_c1, col_c2 = st.columns(2)
            with col_c1:
                ciclo_a_lbl = st.selectbox("Ciclo A (anterior)", aud_nomes2, key="comp_a")
                ciclo_a_id  = df_all2["id"].iloc[aud_nomes2.index(ciclo_a_lbl)]
            with col_c2:
                default_b = min(1, len(aud_nomes2)-1)
                ciclo_b_lbl = st.selectbox("Ciclo B (atual)", aud_nomes2,
                                           index=default_b, key="comp_b")
                ciclo_b_id  = df_all2["id"].iloc[aud_nomes2.index(ciclo_b_lbl)]

            if st.button("🔍 Comparar", key="btn_comp"):
                with st.spinner("Calculando comparativo…"):
                    df_comp = db_module.comparativo_ciclos(int(ciclo_a_id), int(ciclo_b_id))

                if df_comp.empty:
                    st.warning("Sem dados para comparar. Verifique se os ciclos possuem avaliações salvas.")
                else:
                    # Métricas de resumo
                    melhorou = (df_comp["delta"] > 0).sum()
                    piorou   = (df_comp["delta"] < 0).sum()
                    igual    = (df_comp["delta"] == 0).sum()
                    sem_dado = df_comp["delta"].isna().sum()

                    mc1, mc2, mc3, mc4 = st.columns(4)
                    mc1.metric("⬆️ Melhoraram", int(melhorou))
                    mc2.metric("⬇️ Pioraram",   int(piorou))
                    mc3.metric("➡️ Sem mudança", int(igual))
                    mc4.metric("— Sem nota",    int(sem_dado))

                    st.markdown("---")
                    # Tabela colorida
                    def color_delta(v):
                        if pd.isna(v) or v == "—": return ""
                        if "Melhora" in str(v):  return "color: #388E3C; font-weight:bold"
                        if "Piora"   in str(v):  return "color: #D32F2F; font-weight:bold"
                        return "color: #888"

                    cols_show = ["pratica_num","pratica_nome","subitem_idx","subitem_nome",
                                 "nota_a","nota_b","delta","tendencia"]
                    df_disp = df_comp[cols_show].copy()
                    df_disp.columns = ["Prática","Nome Prática","Sub","Nome Sub",
                                       "Nota Ciclo A","Nota Ciclo B","Delta","Tendência"]
                    styled = df_disp.style.applymap(
                        color_delta, subset=["Tendência"]
                    )
                    st.dataframe(styled, use_container_width=True, hide_index=True)

                    # Gráfico de evolução (barras agrupadas)
                    try:
                        import altair as alt
                        df_melt = df_comp.dropna(subset=["nota_a","nota_b"]).copy()
                        df_melt["label"] = df_melt.apply(
                            lambda r: f"{int(r.pratica_num)}.{int(r.subitem_idx)}", axis=1)
                        df_long = pd.melt(
                            df_melt[["label","nota_a","nota_b"]],
                            id_vars="label", value_vars=["nota_a","nota_b"],
                            var_name="Ciclo", value_name="Nota"
                        )
                        df_long["Ciclo"] = df_long["Ciclo"].map(
                            {"nota_a": ciclo_a_lbl[:30], "nota_b": ciclo_b_lbl[:30]})
                        chart = alt.Chart(df_long).mark_bar().encode(
                            x=alt.X("label:N", title="Subitem", sort=None),
                            y=alt.Y("Nota:Q", scale=alt.Scale(domain=[0,4])),
                            color=alt.Color("Ciclo:N",
                                scale=alt.Scale(range=["#90CAF9","#1565C0"])),
                            xOffset="Ciclo:N",
                            tooltip=["label","Ciclo","Nota"]
                        ).properties(
                            title="Comparativo de notas por subitem",
                            height=320
                        )
                        st.altair_chart(chart, use_container_width=True)
                    except Exception:
                        pass  # altair não disponível — pular gráfico

    # ════════════════════════════════════════════════════════════
    # TAB 3 — Backups
    # ════════════════════════════════════════════════════════════
    with tab_backup:
        st.subheader("💾 Backups Automáticos")

        bkp_dir = db_module.BACKUP_DIR
        st.info(f"📂 Pasta de backups: `{bkp_dir}`")

        # Fazer backup manual
        col_bk1, col_bk2 = st.columns([1,3])
        with col_bk1:
            if st.button("💾 Fazer Backup Agora", use_container_width=True):
                bk = db_module.fazer_backup(motivo="manual")
                if bk:
                    st.success(f"✅ Backup criado: `{bk.name}`")
                else:
                    st.warning("Banco ainda não existe (nenhuma auditoria criada).")

        st.markdown("---")
        st.markdown("**Backups disponíveis** (mais recente primeiro):")

        backups = db_module.listar_backups()
        if not backups:
            st.info("Nenhum backup encontrado.")
        else:
            for bk in backups:
                bc1, bc2, bc3, bc4 = st.columns([3,2,1,1])
                bc1.markdown(f"📄 `{bk['nome']}`")
                bc2.markdown(f"🕐 {bk['data']}")
                bc3.markdown(f"**{bk['tamanho_kb']} KB**")
                if bc4.button("🔄 Restaurar", key=f"restore_{bk['nome']}"):
                    ok = db_module.restaurar_backup(bk["path"])
                    if ok:
                        st.success(f"✅ Banco restaurado de `{bk['nome']}`! Recarregue a página.")
                        st.cache_data.clear()
                        st.rerun()
                    else:
                        st.error("Falha ao restaurar.")

        st.markdown("---")
        st.caption(f"🔄 Política: backup automático a cada sessão e antes de exclusões. Mantém os últimos {db_module.MAX_BACKUPS} arquivos.")

    # ════════════════════════════════════════════════════════════
    # TAB 4 — Exportar / Importar
    # ════════════════════════════════════════════════════════════
    with tab_io:
        st.subheader("📦 Exportar / Importar Banco Completo")
        st.markdown("Útil para mover os dados entre computadores ou fazer backup externo.")

        col_exp, col_imp = st.columns(2)

        with col_exp:
            st.markdown("#### 📤 Exportar")
            st.markdown("Gera uma cópia portátil do banco com **todas** as auditorias e avaliações.")
            if st.button("📤 Gerar Arquivo de Exportação", use_container_width=True):
                ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
                dst = db_module.DADOS_DIR / f"export_completo_{ts}.db"
                exp = db_module.exportar_banco(dst)
                # Leitura binária para download
                with open(exp, "rb") as ef:
                    st.download_button(
                        label="⬇️ Baixar banco exportado",
                        data=ef.read(),
                        file_name=exp.name,
                        mime="application/octet-stream",
                        use_container_width=True
                    )

        with col_imp:
            st.markdown("#### 📥 Importar")
            st.warning("⚠️ Substituirá o banco atual. Um backup é criado automaticamente antes.")
            up_file = st.file_uploader("Selecione o arquivo .db exportado",
                                       type=["db"], key="import_db_file")
            if up_file and st.button("📥 Importar e Substituir", key="btn_importar"):
                # Salvar upload temporário
                tmp = db_module.DADOS_DIR / "import_tmp.db"
                with open(tmp, "wb") as tf:
                    tf.write(up_file.read())
                ok = db_module.importar_banco(tmp, modo="substituir")
                tmp.unlink(missing_ok=True)
                if ok:
                    st.success("✅ Banco importado com sucesso! Recarregue a página.")
                    st.cache_data.clear()
                    st.rerun()
                else:
                    st.error("Falha ao importar. Verifique o arquivo.")

    # ════════════════════════════════════════════════════════════
    # TAB 5 — Audit Log
    # ════════════════════════════════════════════════════════════
    with tab_log:
        st.subheader("🔍 Audit Log — Histórico de Alterações")
        st.markdown("Rastreia cada mudança de decisão ou nota registrada no sistema.")

        df_all3 = listar_auditorias()
        if df_all3.empty:
            st.info("Sem auditorias cadastradas.")
        else:
            aud_nomes3 = [f"[{r.id}] {r.unidade} / {r.area} — {r.ciclo}"
                         for r in df_all3.itertuples()]
            sel_log = st.selectbox("Auditoria", aud_nomes3, key="sel_log_aud")
            sel_log_id = df_all3["id"].iloc[aud_nomes3.index(sel_log)]

            limit_log = st.slider("Últimas N alterações", 20, 500, 100, step=20)
            df_log = db_module.carregar_log(int(sel_log_id), limit=limit_log)

            if df_log.empty:
                st.info("Nenhuma alteração registrada ainda para esta auditoria.")
            else:
                # Formatar timestamp
                df_log["timestamp"] = pd.to_datetime(
                    df_log["timestamp"], errors="coerce"
                ).dt.strftime("%d/%m/%Y %H:%M:%S")
                df_log.columns = [c.replace("_"," ").title() for c in df_log.columns]
                st.dataframe(df_log, use_container_width=True, hide_index=True)
                st.caption(f"Mostrando as últimas {len(df_log)} alterações.")

    # ── ABA: APRENDIZADOS IA ─────────────────────────────────────────────────
    with tab_learn:
        st.subheader("🧠 Aprendizados Acumulados pela IA")
        st.caption(
            "Cada vez que o auditor corrige a decisão da IA via revisão colaborativa, "
            "o caso é registrado aqui como exemplo. Nas próximas análises do mesmo subitem, "
            "a IA receberá estes casos como referência (few-shot learning)."
        )
        try:
            aprendizados = db_module.listar_todos_aprendizados(limit=200)
        except Exception:
            aprendizados = []

        if not aprendizados:
            st.info("Nenhum aprendizado registrado ainda. Realize revisões colaborativas para que a IA aprenda com seu julgamento de campo.")
        else:
            df_ap = pd.DataFrame(aprendizados)
            col_a1, col_a2, col_a3 = st.columns(3)
            with col_a1:
                st.metric("Total de casos", len(df_ap))
            with col_a2:
                n_ins = len(df_ap[df_ap["decisao_consenso"]=="insuficiente"]) if "decisao_consenso" in df_ap.columns else 0
                st.metric("Evidência Insuficiente", n_ins)
            with col_a3:
                n_red = (len(df_ap[df_ap["nota_consenso"] < df_ap["nota_ia_inicial"]])
                         if all(c in df_ap.columns for c in ["nota_consenso","nota_ia_inicial"]) else 0)
                st.metric("Notas reduzidas pelo auditor", n_red)
            st.markdown("---")
            if "pratica_nome" in df_ap.columns:
                praticas_ap = ["Todas"] + sorted(df_ap["pratica_nome"].dropna().unique().tolist())
                filtro_prat = st.selectbox("Filtrar por prática:", praticas_ap, key="ap_filtro_prat")
