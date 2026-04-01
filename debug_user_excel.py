import openpyxl
import re
from pathlib import Path

# The path provided by the user
file_path = r"C:\Users\Duda PC\OneDrive\Documentos\Automateasy\Auditoria\Self Assessment\Juiz de Fora\Assessment Automação 2026_AMJF_Redução.xlsx"

if not Path(file_path).exists():
    print(f"File NOT FOUND at {file_path}")
    exit()

def _safe_int(v):
    if v is None: return None
    try:
        if isinstance(v, float):
            if v != v: return None
            return int(v)
        if isinstance(v, str):
            v = v.strip().split('.')[0]
            if not v: return None
            return int(v)
        return int(v)
    except: return None

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet_name = next((n for n in wb.sheetnames if 'ROAD MAP' in n and 'Trefila' not in n), wb.sheetnames[0])
    print(f"Using sheet: {sheet_name}")
    ws = wb[sheet_name]
    
    current_p_num = None
    current_s_offset = 0
    results = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if not row or all(v is None for v in row): continue
        
        first_col = row[0]
        col2_val = str(row[2] or "").strip()
        
        # 1. Header/Footer Skip
        if isinstance(first_col, str) and ("N°" in first_col or len(first_col) > 10):
            current_p_num = None # Stop current practice context on header
            continue
        if not col2_val or col2_val.upper() == "EVIDÊNCIA":
            continue

        # 2. Practice Detection
        if isinstance(first_col, int):
            current_p_num = first_col
            current_s_offset = 0
            # This row is ALSO subitem 1
        elif first_col is None and current_p_num is not None:
            # Persistent subitem
            current_s_offset += 1
        elif isinstance(first_col, str):
            # Fallback for "1.1" regex
            m = re.match(r'^(\d+)\.(\d+)', first_col.strip())
            if m:
                current_p_num = int(m.group(1))
                current_s_offset = int(m.group(2)) - 1
            else:
                # Standalone numeric string "1"
                m_p = re.match(r'^(\d+)\s*$', first_col.strip())
                if m_p:
                    current_p_num = int(m_p.group(1))
                    current_s_offset = 0
                else:
                    continue # Skip other noise
        else:
            continue

        # 3. Score Extraction
        final_nota = None
        for col_idx in [8, 9, 7]:
            if len(row) > col_idx:
                val = _safe_int(row[col_idx])
                if val is not None:
                    final_nota = val
                    break
        
        results.append({
            'row': i+1,
            'p': current_p_num,
            's': current_s_offset + 1,
            'desc': col2_val[:50],
            'nota': final_nota
        })

    print(f"\nFound {len(results)} items:")
    for res in results[:20]:
        print(f"Row {res['row']}: Practice {res['p']} Subitem {res['s']} | Nota={res['nota']} | {res['desc']}...")
    
    # Check total for Practice 1
    p1 = [r for r in results if r['p'] == 1]
    print(f"\nPractice 1 total items: {len(p1)}")
    if p1:
        print(f"Practice 1 last subitem: {p1[-1]['s']}")
    
except Exception as e:
    import traceback
    traceback.print_exc()
