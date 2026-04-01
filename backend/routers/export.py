"""
routers/export.py — Excel export and new audit creation.
"""
from fastapi import APIRouter, HTTPException, UploadFile, File, Form, BackgroundTasks, Depends
from fastapi.responses import FileResponse
from pydantic import BaseModel
from backend.db import get_db, get_auditoria, carregar_avaliacoes, _safe_int, USE_POSTGRES
from backend.auth import get_current_user
from datetime import datetime
from pathlib import Path
import tempfile
import json
import sys
import os
import shutil
import zipfile
import logging

log = logging.getLogger("auditoria_export")

_parent = str(Path(__file__).resolve().parent.parent.parent)
if _parent not in sys.path:
    sys.path.insert(0, _parent)

router = APIRouter(prefix="/api", tags=["export"])

# Diretório base para os uploads na nuvem
IS_VERCEL = bool(os.environ.get("VERCEL") or os.environ.get("VERCEL_ENV"))
if IS_VERCEL:
    UPLOAD_DIR = Path("/tmp/AuditoriaTA/uploads")
else:
    UPLOAD_DIR = Path("data/uploads")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

def process_heavy_files(audit_id: int, assessment_url: str, evidence_url: str, assessment_file_bytes: bytes, evidence_zip_bytes: bytes):
    """Background task to handle heavy file processing without timing out the request."""
    audit_dir = UPLOAD_DIR / str(audit_id)
    audit_dir.mkdir(parents=True, exist_ok=True)
    
    assessment_path = ""
    evidence_path = ""
    
    # 1. Process Assessment
    if assessment_file_bytes:
        file_path = audit_dir / f"assessment_{audit_id}.xlsx"
        with open(file_path, "wb") as buffer:
            buffer.write(assessment_file_bytes)
        assessment_path = str(file_path.absolute())
    elif assessment_url:
        if os.path.isabs(assessment_url) and os.path.exists(assessment_url):
            file_path = audit_dir / f"assessment_{audit_id}.xlsx"
            shutil.copy2(assessment_url, file_path) # Usar copy2 em vez de move para evitar problemas de permissão
            assessment_path = str(file_path.absolute())
        else:
            import urllib.request
            try:
                file_path = audit_dir / f"assessment_{audit_id}.xlsx"
                urllib.request.urlretrieve(assessment_url, file_path)
                assessment_path = str(file_path.absolute())
            except Exception as e:
                log.error(f"Failed to download assessment: {e}")

    # 2. Process Evidence ZIP
    if evidence_zip_bytes:
        zip_path = audit_dir / f"evidences_{audit_id}.zip"
        with open(zip_path, "wb") as buffer:
            buffer.write(evidence_zip_bytes)
        
        extract_dir = audit_dir / "evidences"
        try:
            from backend.routers.evidencias import extract_zip_robustly
            extract_zip_robustly(zip_path, extract_dir)
            evidence_path = str(extract_dir.absolute())
            # Cleanup ZIP after extraction to save space
            if zip_path.exists():
                os.remove(zip_path)
                log.info(f"Removed temporary ZIP: {zip_path}")
        except Exception as e:
            log.error(f"Failed to extract: {e}")
            evidence_path = str(extract_dir.absolute())
    elif evidence_url:
        if os.path.isabs(evidence_url) and os.path.exists(evidence_url):
            zip_path = audit_dir / f"evidences_{audit_id}.zip"
            shutil.copy2(evidence_url, zip_path)
            extract_dir = audit_dir / "evidences"
            try:
                from backend.routers.evidencias import extract_zip_robustly
                extract_zip_robustly(zip_path, extract_dir)
                evidence_path = str(extract_dir.absolute())
            except Exception as e:
                log.error(f"Failed to extract local URL zip: {e}")
                evidence_path = str(extract_dir.absolute())
        else:
            import urllib.request
            try:
                zip_path = audit_dir / f"evidences_{audit_id}.zip"
                urllib.request.urlretrieve(evidence_url, zip_path)
                extract_dir = audit_dir / "evidences"
                from backend.routers.evidencias import extract_zip_robustly
                extract_zip_robustly(zip_path, extract_dir)
                evidence_path = str(extract_dir.absolute())
                # Cleanup ZIP after extraction to save space
                if zip_path.exists():
                    os.remove(zip_path)
                    log.info(f"Removed downloaded ZIP: {zip_path}")
            except Exception as e:
                log.error(f"Failed to download/extract URL zip: {e}")

    # 3. Build Evidence Map
    evidence_map = {}
    if evidence_path and Path(evidence_path).is_dir():
        try:
            from backend.routers.evidencias import _get_or_build_evidence_map
            evidence_map = _get_or_build_evidence_map(evidence_path, refresh=True)
            # Standardize keys for JSON (tuple -> "P.S")
            evidence_map = {f"{p}.{s}": files for (p, s), files in evidence_map.items()}
        except Exception as e:
            log.error(f"Map failed: {e}")

    # 4. Final DB Update (Paths and Evidence Map)
    with get_db() as conn:
        q = "UPDATE auditorias SET assessment_file_path=?, evidence_folder_path=?, evidence_map=?, evidence_zip_url=? WHERE id=?"
        if USE_POSTGRES: q = q.replace("?", "%s")
        conn.execute(q, (assessment_path, evidence_path, json.dumps(evidence_map), evidence_url or "", audit_id))
        conn.commit()

    # 5. Populate Subitems (Baseline)
    try:
        from checklist_po_aut_002 import CHECKLIST
        from criterios_oficiais import CRITERIOS
        with get_db() as conn:
            for key, info in CHECKLIST.items():
                p_num, s_idx = key
                p_nome = "Rotinas de TA" # Default
                
                # Get official names if available
                crit_info = CRITERIOS.get(key, {})
                s_nome = crit_info.get('subitem') or f"Subitem {s_idx+1}"
                p_nome_oficial = crit_info.get('pratica')
                
                # Fallback mapping for practice names
                pratica_nomes = {1:"Rotinas de TA", 2:"Sobressalentes", 3:"Mapa de Ativos", 4:"Conhecimento", 5:"Infraestrutura", 6:"Riscos", 7:"TI", 8:"Software/Hardware", 9:"Cyber"}
                p_nome = p_nome_oficial or pratica_nomes.get(p_num, p_nome)

                q_ins = "INSERT OR IGNORE INTO avaliacoes (auditoria_id, pratica_num, pratica_nome, subitem_idx, subitem_nome, evidencia_descricao, decisao) VALUES (?,?,?,?,?,?,'pendente')"
                if USE_POSTGRES: 
                    q_ins = "INSERT INTO avaliacoes (auditoria_id, pratica_num, pratica_nome, subitem_idx, subitem_nome, evidencia_descricao, decisao) VALUES (%s,%s,%s,%s,%s,%s,'pendente') ON CONFLICT DO NOTHING"
                
                conn.execute(q_ins, (audit_id, p_num, p_nome, s_idx, s_nome, " \n".join(info.get("verificar", []))))
            conn.commit()
    except Exception as e:
        log.error(f"Background baseline population failed: {e}")

    # 6. Parse Excel for SA Notes (Optional)
    if assessment_path and Path(assessment_path).exists():
        try:
            import openpyxl
            wb = openpyxl.load_workbook(assessment_path, data_only=True)
            ws = wb.active
            _parse_assessment_sheet(ws, audit_id)
        except Exception as e:
            log.error(f"Background Excel sync failed for audit {audit_id}: {e}")

def _parse_assessment_sheet(ws, audit_id: int):
    """
    Robustly parses the assessment worksheet to extract SA scores.
    Supports:
    1. Standard Checklist (scores in Col I/J, descriptions in Col C)
    2. Integrated Report (scores in Col B, descriptions in Col A)
    """
    import re
    from backend.db import get_db, _safe_int
    
    current_p_num = None
    updates = []
    is_integrated = False
    
    # Headers to skip
    SKIP_KEYWORDS = {
        "EVIDÊNCIA", "SUBITEM", "DESCRIÇÃO", "PRÁTICA", "REQUISITO", 
        "EVIDENCIAS", "N°", "NÃO TEM PRÁTICA", "PONTUAÇÃO", 
        "STATUS DA NOTA", "TIPO DE NÃO CONFORMIDADE"
    }

    log.info(f"Refined parsing started for audit {audit_id}")

    # 1. Detect format by inspecting the first few rows
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        if not row: continue
        row_str = " ".join(str(c).upper() for c in row if c)
        if "STATUS DA NOTA" in row_str or "COMENTÁRIOS ADICIONAIS" in row_str:
            is_integrated = True
            log.info(f"Audit {audit_id}: Detected 'Integrated Report' format.")
            break

    # 2. Parse rows
    current_p_num = None
    current_s_offset = 0

    for i, row_cells in enumerate(ws.iter_rows(min_row=1, values_only=True)):
        if not row_cells or all(v is None for v in row_cells):
            continue
        
        first_col = row_cells[0]
        col2_val = str(row_cells[2] or "").strip() if len(row_cells) > 2 else ""
        
        # 1. Header/Footer Skip
        if isinstance(first_col, str) and ("N°" in first_col or len(first_col) > 10):
            current_p_num = None
            continue
        if not col2_val or col2_val.upper() == "EVIDÊNCIA":
            continue

        # 2. Practice & Subitem Detection
        p_num = None
        s_idx = None

        # 2a. Explicit Match (1.1 or 6.A)
        explicit_match = None
        if isinstance(first_col, str):
            explicit_match = re.match(r'^(\d+)\.([A-Za-z0-9]+)', first_col.strip())
        
        if explicit_match:
            p_num = int(explicit_match.group(1))
            s_part = explicit_match.group(2)
            if s_part.isdigit():
                s_idx = int(s_part) - 1
            else:
                # Convert A->0, B->1, etc.
                s_idx = ord(s_part.upper()) - ord('A')
            
            current_p_num = p_num
            current_s_offset = s_idx
        # 2b. Practice Header (int)
        elif isinstance(first_col, int):
            p_num = first_col
            s_idx = 0
            current_p_num = p_num
            current_s_offset = 0
        # 2c. Practice Header (string "1")
        elif isinstance(first_col, str):
            m_p = re.match(r'^(\d+)\s*$', first_col.strip())
            m_ip = re.match(r'^(\d+)\s*[-–]', first_col.strip())
            m = m_p or m_ip
            if m and len(first_col) < 50:
                p_num = int(m.group(1))
                s_idx = 0
                current_p_num = p_num
                current_s_offset = 0
        # 2d. Positional
        elif first_col is None and current_p_num is not None:
            current_s_offset += 1
            p_num = current_p_num
            s_idx = current_s_offset
        
        if p_num is None or s_idx is None:
            continue

        # 3. Score Extraction
        nota_sa = None
        
        # Priority columns based on format
        check_cols = [8, 9, 7] # Default: I, J, H
        if is_integrated:
            check_cols = [1, 8, 9, 7] # Priority to Column B for Integrated Report
            
        for col_idx in check_cols:
            if len(row_cells) > col_idx:
                val = _safe_int(row_cells[col_idx])
                if val is not None:
                    nota_sa = val
                    break

        if nota_sa is not None:
            updates.append((nota_sa, audit_id, p_num, s_idx))
            log.debug(f"Row {i+1}: Found P{p_num} S{s_idx+1} = {nota_sa}")

    # 3. Apply updates
    if updates:
        with get_db() as conn:
            for params in updates:
                conn.execute("""
                    UPDATE avaliacoes
                    SET nota_self_assessment=?
                    WHERE auditoria_id=? AND pratica_num=? AND subitem_idx=?
                """, params)
            conn.commit()
        log.info(f"Audit {audit_id}: Finished syncing {len(updates)} subitems from Excel.")
    
    return len(updates)

@router.post("/auditorias")
async def criar_auditoria(
    background_tasks: BackgroundTasks,
    unidade: str = Form(...),
    area: str = Form(...),
    ciclo: str = Form(...),
    openai_api_key: str = Form(""),
    observacoes: str = Form(""),
    modo_analise: str = Form("completo"),
    assessment_file: UploadFile = File(None),
    assessment_url: str = Form(""),
    evidence_zip: UploadFile = File(None),
    evidence_url: str = Form(""),
    current_user: str = Depends(get_current_user)
):
    """Create audit skeleton and offload processing to background task."""
    now = datetime.now().isoformat()
    
    with get_db() as conn:
        q = """
            INSERT INTO auditorias
                (unidade, area, ciclo, data_criacao, data_atualizacao,
                 status, assessment_file_path, evidence_folder_path,
                 openai_api_key, observacoes, modo_analise, ai_provider, ai_base_url, auditado_por)
            VALUES (?,?,?,?,?,'em_andamento','','',?,?,?,'','',?)
        """
        if USE_POSTGRES: q = q.replace("?", "%s")
        conn.execute(q, (unidade, area, ciclo, now, now, openai_api_key, observacoes, modo_analise, current_user))
        conn.commit()
        
        row = conn.execute("SELECT id FROM auditorias ORDER BY id DESC LIMIT 1").fetchone()
        if not row:
            raise HTTPException(status_code=500, detail="Erro ao reservar ID da auditoria")
        audit_id = row["id"]

    # Read bytes to pass to background task (Uvicorn handles large files well, but we must read them before request ends)
    as_bytes = await assessment_file.read() if assessment_file else None
    ev_bytes = await evidence_zip.read() if evidence_zip else None
    
    background_tasks.add_task(
        process_heavy_files, 
        audit_id, assessment_url, evidence_url, as_bytes, ev_bytes
    )

    return {
        "ok": True, 
        "id": audit_id, 
        "message": "Auditoria criada com sucesso! Os arquivos estão sendo processados em segundo plano e as evidências aparecerão em instantes."
    }

@router.post("/auditorias/{auditoria_id}/importar-assessment")
def importar_assessment(auditoria_id: int, assessment_path: str = ""):
    """Manual trigger to re-import assessment Excel."""
    aud = get_auditoria(auditoria_id)
    if not aud:
        raise HTTPException(status_code=404, detail="Auditoria não encontrada")

    path = assessment_path or aud.get("assessment_file_path", "")
    if not path or not Path(path).exists():
        raise HTTPException(status_code=400, detail="Arquivo de assessment não encontrado")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        imported = _parse_assessment_sheet(ws, auditoria_id)
        return {"ok": True, "imported": imported}
    except Exception as e:
        log.error(f"Manual import failed for audit {auditoria_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao importar: {str(e)}")

@router.get("/auditorias/{auditoria_id}/exportar-excel")
def exportar_excel(auditoria_id: int):
    """Export audit results exactly matching the 'Integrated Report' (Auditoria_teste.xlsx) format."""
    aud = get_auditoria(auditoria_id)
    if not aud: raise HTTPException(status_code=404, detail="Não encontrado")
    
    df = carregar_avaliacoes(auditoria_id)
    df.sort(key=lambda x: (x.get('pratica_num') or 0, x.get('subitem_idx') or 0))
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Análise de Auditoria"
    
    # Precise Style Definitions
    title_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    practice_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    red_status_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_status_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    yellow_nc_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    orange_nc_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    grey_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
    
    white_bold_font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
    navy_bold_font = Font(name='Arial', size=10, bold=True, color="002060")
    black_bold_font = Font(name='Arial', size=10, bold=True, color="000000")
    black_normal_font = Font(name='Arial', size=10, color="000000")
    title_font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'), 
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_centered_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 1. Main Title (Row 1)
    title_text = f"FORMULÁRIO PARA ANÁLISE DAS NOTAS DO SELF ASSESSMENT DA {aud['area'].upper()} DE {aud['unidade'].upper()}"
    ws.merge_cells('A1:F1')
    cell_a1 = ws['A1']
    cell_a1.value = title_text
    cell_a1.fill = title_fill
    cell_a1.font = title_font
    cell_a1.alignment = center_align
    ws.row_dimensions[1].height = 40

    # 2. Headers (Row 2)
    headers = ['PRÁTICAS', 'Nota', 'Status da Nota', 'Tipo de Não Conformidade', 'Descrição da Não Conformidade', 'Comentários Adicionais']
    ws.append(headers)
    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = black_bold_font
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[2].height = 30

    # 3. Content
    current_p_num = None
    row_idx = 3
    import re
    
    for av in df:
        p_num = av.get('pratica_num')
        p_nome = str(av.get('pratica_nome', '')).strip()
        
        # New Practice Header Row
        if p_num != current_p_num:
            current_p_num = p_num
            ws.merge_cells(f'A{row_idx}:F{row_idx}')
            
            # Clean p_nome to avoid double numbering (e.g., "1 - 1 - ROTINAS")
            p_label = p_nome.upper()
            # If it already starts with "N - " or "N – " or just "N", clean it
            p_label = re.sub(r'^\d+[\s\-\–]*', '', p_label).strip()
            p_full_label = f"{p_num} - {p_label}"
            
            cell = ws.cell(row=row_idx, column=1, value=p_full_label)
            cell.fill = practice_fill
            cell.font = navy_bold_font
            cell.alignment = Alignment(vertical="center", indent=1)
            for c in range(1, 7):
                ws.cell(row=row_idx, column=c).border = thin_border
            ws.row_dimensions[row_idx].height = 25
            row_idx += 1
        
        # Subitem Data Row
        nota_sa = av.get('nota_self_assessment')
        nota_f = av.get('nota_final')
        
        # Determinando o Status da Nota e Tipo de NC
        status = "Permanece"
        if nota_f is not None and nota_sa is not None:
            if nota_f < nota_sa: status = "Diminui"
            elif nota_f > nota_sa: status = "Aumenta"
            
        decisao = av.get('decisao', 'pendente')
        tipo_nc = "Não se aplica"
        if status == "Diminui":
            if decisao == 'inexistente':
                tipo_nc = "Evidências inexistente"
            else:
                tipo_nc = "Evidências insuficiente"
            
        # Subitem Description refinement
        # User wants only the short name without the number (e.g. "Backup..." instead of "1.1 - Backup...")
        s_nome = str(av.get('subitem_nome', '')).strip()
        
        # Strip leading "1.1 - " or similar
        final_desc = re.sub(r'^\d+\.\d+[\s\-\–]*', '', s_nome).strip()
        
        vals = [
            final_desc,
            nota_sa if nota_sa is not None else '',
            status,
            tipo_nc,
            av.get('descricao_nc', '') or '',
            av.get('comentarios', '') or ''
        ]
        
        ws.append(vals)
        for ci, cell in enumerate(ws[row_idx], 1):
            cell.font = black_normal_font
            cell.border = thin_border
            cell.alignment = left_centered_align if ci in [1, 5, 6] else center_align
            
            # Conditional Styling for Status (Column C)
            if ci == 3:
                if vals[2] == "Diminui":
                    cell.fill = red_status_fill
                    cell.font = white_bold_font
                elif vals[2] in ["Permanece", "Mantém", "Manter", "Aumenta", "Melhora"]:
                    cell.fill = green_status_fill
                    cell.font = white_bold_font
            
            # Conditional Styling for Tipo NC (Column D)
            if ci == 4:
                if vals[3] == "Evidências inexistente":
                    cell.fill = grey_fill
                elif vals[2] == "Diminui":
                    cell.fill = yellow_nc_fill
                else:
                    cell.fill = orange_nc_fill
                
        row_idx += 1

    # Column Widths
    widths = [67, 17, 18, 28, 93, 107]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()
    
    return FileResponse(tmp.name, filename=f"Relatorio_{auditoria_id}_{aud['unidade']}_{aud['area']}.xlsx")

@router.get("/unidades-areas")
def get_unidades_areas():
    return {
        "Juiz de Fora": ["Utilidades", "Aciaria", "Laminação", "Redução"],
        "Piracicaba": ["Aciaria", "Utilidades", "Laminação"],
        "Sitrel": ["Laminação"],
        "João Monlevade": ["Aciaria", "Laminação TL2", "Laminação TL3", "Alto Forno", "Utilidades"],
        "Trefilarias": ["São Paulo", "Sabará", "Resende", "Juiz de Fora"],
        "Mina do Andrade": ["Mineração"]
    }
    
